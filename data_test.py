

from datetime import datetime
from download import excel_to_pandas
from default import URL
from web import to_pdf
import pandas as pd 
from default import COLUMNS, ASK, REPORT_COLS, STATIC_COLS, CHANGE_YESNO, outpath, SAME_COLUMNS
import os 
import sys
import movecolumn as mc 
import numpy as np
from operator import itemgetter
import math 
from utils.utils import getIndexes
import string
import openpyxl
import warnings
warnings.filterwarnings('ignore')

root = os.getcwd()
report_cols_list = [REPORT_COLS[key] for key in list(REPORT_COLS.keys())]
alphabet = dict.fromkeys(string.ascii_uppercase, 0)
alphabet = dict(zip(range(len(list(alphabet.keys()))), list(alphabet.keys())))


def to_excel(df, name):
    name = os.path.join(outpath, name)
    df.to_excel(name, index=False)







def generate_pdf():
    today = str(datetime.today().date()).replace('-', '_')
    db_filename = 'db_{today}.xlsx'.format(today=today)
    df = excel_to_pandas(URL, 'data/' + db_filename)
    return df

    

def report(df):
    mode ='w'
    prefix = ''
    today = str(datetime.today().date()).replace('-', '_')
    report_filename = 'report_analysis_{today}.xlsx'.format(today=today)
    # df=pd.read_excel(f'out_test\\db_{today}.xlsx')

    df['_submission_time'] = pd.to_datetime(df['_submission_time']).dt.date
    mask_eff = ~pd.isnull(df['7.1. Мамлакатдаги умумий вазиятдан (ҳолатдан) қониқиш даражангизни 7 баллик шкалада баҳоланг?'])
    df = df[mask_eff]
    df_orig = df.copy()
    df.rename(columns={'Яшаш ҳудудингиз:': 'region', 'Яшаш тумани:': 'district', '_id': 'count'}, inplace=True)
    df = df[['region', 'district', 'count', '_submission_time']]

    
    dates = df_orig['_submission_time'].unique()
    df_out = pd.DataFrame()
    for i, date in enumerate(sorted(dates)):
        mask = df['_submission_time'] == date
        df_temp = df[mask]
        df_temp.drop('_submission_time', axis=1, inplace=True)
        df_temp = df_temp.groupby(by=['district'], as_index=False).count()
        df_temp.rename(columns={'region':'Ҳудуд',  'district': 'Туман', 'count': date}, inplace=True)
        total = df_temp[date].sum()
        df_temp.drop('Ҳудуд', axis=1, inplace=True)
        
        
        if i == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='out_tester', on=['Туман'])
    
    
    total_col = df_out[dates.tolist()].sum(axis=1)
    df_out['Жами']= total_col

    # sys.exit()



    # df_total = pd.DataFrame(columns=df_out.columns, data=[['Жами:', np.nan, np.nan, np.nan, np.nan, np.nan, np.nan , total]])
    df_total = df_out[dates].sum()
    total_all = sum(df_total.values.tolist())
    df_total = pd.DataFrame(columns=['Туман'] +df_total.index.values.tolist(), data=[['Жами:']+df_total[dates].values.tolist()])
    
    df_out = pd.concat([df_out, df_total], axis=0)
    df_out.reset_index(drop=True, inplace=True)
    df_out.at[df_out.index.values.tolist()[-1], 'Жами'] = total_all
    with pd.ExcelWriter(f'out_test/report/{report_filename}', mode=mode, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}districts')

    # print(df_out)
    ############################################################################ Вилоятлар    
    df_orig.rename(columns={'Яшаш ҳудудингиз:': 'region', 'Яшаш тумани:': 'district', '_id': 'count'}, inplace=True)
    df_orig = df_orig[['region', 'district', 'count', '_submission_time']]

    # df_orig = df_orig.groupby(by=['region'], as_index=False).count()[['region', 'count']]
    # df_orig.rename(columns={'region':'Ҳудуд', 'count': 'Уй хўжаликлар сони'}, inplace=True)
    df_out = pd.DataFrame()
   
    for k, date in enumerate(dates):
        mask = df_orig['_submission_time'] == date
        df_temp = df_orig[mask]
        df_temp = df_temp.groupby(by=['region'], as_index=False).count()[['region', 'count']]
        total = df_temp['count'].sum()
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        df_total = pd.DataFrame(columns=df_temp.columns, data=[['Жами:', total]])
        
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        
        if k == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='out_tester', on=['Ҳудуд'])
    df_out['Жами'] = df_out[dates].sum(axis=0)
    total_row = df_out.sum()[dates]
    total_row["Ҳудуд"] = 'Жами:'
    total_row = total_row.reset_index().T.reset_index(drop=True)
    total_row.rename(dict(zip(total_row.columns.values.tolist(), total_row.iloc[0].values.tolist())), inplace=True, axis=1)
    total_row.drop(0, axis=0, inplace=True)
    df_out = pd.concat([df_out, total_row], axis=0)
    columns = ["Ҳудуд"] + sorted(dates.tolist()) + ['Жами']
    df_out = df_out[columns]
    df_out['Жами'] = df_out[dates].sum(axis=1)
    with pd.ExcelWriter(f'out_test/report/{report_filename}', mode='a', engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}regions')
    print(f'{datetime.today()} - Starting generating REPORT file...')
    return report_filename


def highlight(x):
    print(x)
    return x

def freq_table(df):
    
    # if 'out_test' not in df_filename:
    #     df_filename = os.path.join(root, 'out_test', df_filename)
    # df = pd.read_excel(df_filename)
    mask_eff = ~pd.isnull(df['7.1. Мамлакатдаги умумий вазиятдан (ҳолатдан) қониқиш даражангизни 7 баллик шкалада баҳоланг?'])
    df = df[mask_eff]
    df_orig = df.copy()
    
    
    today = str(datetime.today().date()).replace('-', '_')

    columns = ['7.1. Мамлакатдаги умумий вазиятдан (ҳолатдан) қониқиш даражангизни 7 баллик шкалада баҳоланг?',
               '7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?']
    columns = ['Яшаш ҳудудингиз:'] + columns

    df = df[columns]
    df.reset_index(drop=True, inplace=True)

    columns = df.columns
    df_out = pd.DataFrame()

    filename_out = f'out_test/freq/ecopol_freq_table_{today}.xlsx'
    df_big = pd.DataFrame()

    town_vil = [ 'Шаҳар', 'Қишлоқ', 'Аёл', 'Эркак', '18-35', '36-49', '50-999','all',]
    for i, c in enumerate(list(reversed(df.columns.values))):

        
        df_temp = pd.DataFrame()
        for j, tv in enumerate(town_vil):
            if tv in ['Шаҳар', 'Қишлоқ']:
                mask = df_orig['1. САВОЛ БЕРМАЙ, Қаерда (Шаҳар ёки қишлоқда) яшашини киритинг?'] == tv
                df = df_orig[mask]
            elif tv in ['18-35', '36-49', '50-999']:
                btm_age, top_age = tv.split('-')
                btm_age, top_age = int(btm_age), int(top_age)
                if btm_age == 18:
                    btm_age = 17
                btm_mask = df_orig['6. Ёшингиз?'] >= btm_age
                top_mask = df_orig['6. Ёшингиз?'] <= top_age
                # mask = ~(btm_mask & top_mask)
                df = df_orig[btm_mask][top_mask]
                df = df.dropna(subset='6. Ёшингиз?', axis=0)
                
            elif tv in ['Аёл', 'Эркак']:
                mask = df_orig['2. САВОЛ БЕРМАЙ, респондентларнинг жинсини аниқланг:'] == tv
                df = df_orig[mask]
            else:
                df = df_orig.copy()
            
            ctab = pd.crosstab(index = df.index, columns = df[c])
            ctab = ctab.sum()

            ctab_pct = (ctab/ctab.values.sum()).round(decimals=4)
            
            tab = pd.concat([ctab, ctab_pct], axis=1)

            cols = ['Frequency', 'Percentage (%)', c]
            tab[c] = tab.index
            df_out = pd.DataFrame(columns=cols, data=tab.values)
            df_out = mc.MoveTo1(df_out, c)

            if not df[c].dtype == 'object':
                df_out['Average score'] = df_out['Percentage (%)']/100 * df_out[c]
                df_out['Average score'] = df_out['Average score'].round(decimals=3)
            else:
                df_out['Average score'] = [np.nan] * len(df_out['Percentage (%)'])
            
            fillna_cols = df_out.columns.values.tolist()
            
            df_out[['Percentage (%)', 'Average score']] = df_out[['Percentage (%)', 'Average score']] * 100

            
            freq_sum = df_out['Frequency'].sum()
            pct_sum =df_out['Percentage (%)'].sum()
            avg_sum = df_out['Average score'].sum()
            if avg_sum == 0:
                avg_sum = np.nan
            df_out['Frequency'] = df_out['Frequency'].astype(float).astype('Int64')
            df_total = pd.DataFrame(columns=df_out.columns, data=[['Total:', freq_sum,\
                                                                pct_sum,
                                                                 avg_sum]])

            df_out = pd.concat([df_out, df_total], axis=0)
            
            

            
            # Add space after tables
            df_out.rename(columns={'Frequency': f'{tv}_Frequency', 'Percentage (%)': f'{tv}_Percentage (%)', 'Average score': f'{tv}_Average score'}, inplace=True)
            if j == 0:
                df_temp = df_out
            else:
                if tv == 'all' or tv in ['Аёл'] or tv == '18-35':
                    space = pd.DataFrame(np.nan, index=list(range(len(df_out))), columns=['A', 'B'])
                    # space = df_out[df_out.columns.values[:2]].fill(0)
                    df_out.reset_index(drop=True, inplace=True)
                    df_out = pd.concat([space, df_out], axis=1)
                
         
                df_temp = pd.merge(df_temp, df_out, how='out_tester', on=c)
                fillna_cols = df_temp.columns.values.tolist()
                fillna_cols = [fc for fc in fillna_cols if fc != 'Average score' and fc != c and fc not in ['A', 'B'] and 'A_' not in fc and 'B_' not in fc]
                df_temp[fillna_cols] = df_temp[fillna_cols].fillna(0)
            

        df_space = pd.DataFrame(columns=df_temp.columns.values.tolist(), data=[[np.nan]*len(df_temp.columns),
                                                                                [np.nan]*len(df_temp.columns)])
        
        ### move the Total row to the last
        col_first = df_temp.columns.values[0]
        mask = df_temp[col_first] == 'Total:'
        df_temp = pd.concat([df_temp[~mask], df_temp[mask]], axis=0)

        ## fill nan values with zeroes
        # df_temp.fillna(0, inplace=True)

        ## add space between question blocks
        df_temp = pd.concat([df_temp, df_space], axis=0)

        

        # columns = [c] + ['Frequency_x', 'Frequency_y', 'Percentage (%)_x',  'Percentage (%)_y', 'Average score_x', 'Average score_y', 'A', 'B', 'Frequency', 'Percentage (%)', 'Average score']
        # print(df_temp.columns.values)
        # sys.exit()
        columns = [c] + ['Шаҳар_Frequency', 'Қишлоқ_Frequency', 'Шаҳар_Percentage (%)',  'Қишлоқ_Percentage (%)', 'Шаҳар_Average score', 'Қишлоқ_Average score', 'A_x', 'B_x', \
                          'Аёл_Frequency', 'Эркак_Frequency', 'Аёл_Percentage (%)', 'Эркак_Percentage (%)', 'Аёл_Average score', 'Эркак_Average score', 'A_y', 'B_y',
                          '18-35_Frequency', '36-49_Frequency', '50-999_Frequency', '18-35_Percentage (%)', '36-49_Percentage (%)', '50-999_Percentage (%)', '18-35_Average score', '36-49_Average score', '50-999_Average score', 'A', 'B',
                          'all_Frequency', 'all_Percentage (%)', 'all_Average score']
        


        

        micolumns = pd.MultiIndex.from_tuples(
                                            [(None, c),
                                             ("Frequency", "Шаҳар"), (None, "Қишлоқ"), 
                                             ("Percentage (%)", "Шаҳар"), (None, "Қишлоқ"),
                                             ("Average score", "Шаҳар"), (None, "Қишлоқ"),
                                             (None, None),
                                             (None, None),
                                             ("Frequency", "Аёл"), (None, "Эркак"), 
                                             ("Percentage (%)", "Аёл"), (None, "Эркак"),
                                             ("Average score", "Аёл"), (None, "Эркак"),
                                             (None, None),
                                             (None, None),
                                             ("Frequency", '18-35'),(None, '36-49'),(None, '50+'), 
                                             ("Percentage (%)", '18-35'),(None, '36-49'),(None, '50+'),
                                             ("Average score", "18-35"),(None, "36-49"),(None, "50+"),
                                             (None, None),
                                             (None, None),
                                             ('Across all data', "Frequency"), 
                                             (None, "Percentage (%)"),
                                             (None, "Average score")],)
        
        # columns = [c] + ['all_Frequency', 'all_Percentage (%)', 'all_Average score']
   
        # micolumns = pd.MultiIndex.from_tuples(
        #                                     [(None, c),
        #                                      ('Across all data', "Frequency"), 
        #                                      (None, "Percentage (%)"),
        #                                      (None, "Average score")
        #                                      ],)

        df_temp = df_temp[columns]
        df_temp = pd.DataFrame(columns=micolumns, data=df_temp.values)

    
        
        
 
        if i == 0:
            df_big = df_temp
        else:
            
            df_temp = df_temp.reset_index(drop=True).T.reset_index().T
            df_temp.rename(columns=dict(zip(df_temp.columns.values.tolist(), df_big.columns.values.tolist())), inplace=True)
            if i == 1:
                columns = df_big.columns.values.tolist()
                df_big = df_big.reset_index(drop=True).T.reset_index().T
                df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), columns)), inplace=True)
            
            df_big = pd.concat([df_temp, df_big], axis=0)

    
        

    new_columns = [np.nan] * len(df_big.columns.values)
    new_columns[0] = 'EcoPol Frequency table'
    new_columns[6] = today.replace('_', '-')
    df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), new_columns)), inplace=True)
    
    with pd.ExcelWriter(filename_out, mode='w', engine='openpyxl') as writer:
        df_big.to_excel(writer, index=False)



    print(f'{datetime.today()} - Starting generating FREQ TABLE file...')
    return filename_out
    

def report_by_region(df):
    
    # today = str(datetime.today().date()).replace('-', '_')
    # df_filename = f'out_test\\db_{today}.xlsx'
    # df = pd.read_excel(df_filename)
    today = str(datetime.today().date()).replace('-', '_')
    out_filename = f'out_test/regional/regional_{today}.xlsx'
    mask_eff = ~pd.isnull(df['7.1. Мамлакатдаги умумий вазиятдан (ҳолатдан) қониқиш даражангизни 7 баллик шкалада баҳоланг?'])
    df = df[mask_eff]
    df = df.replace({'\(БУНИ ЎҚИМАНГ, агар респондент жавоб беришга қийналса шу вариантни белгиланг\) ':'',
                     '\(БУНИ ЎҚИМАНГ, агар респондент жавоб беришга қийналса шу вариантни белгиланг \) ':'',}, regex=True)
    cols = ['7.1. Мамлакатдаги умумий вазиятдан (ҳолатдан) қониқиш даражангизни 7 баллик шкалада баҳоланг?',
            '7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?',
            '11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?',
            '9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?',
            '12. Сизнингча, атрофингиздаги одамлар ҳозир қандай кайфиятда: кўтаринки, хотиржам ёки тушкун (хавотирли)?',
            '13. “Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга…:',
            '16. “Ҳукумат фуқаролар билан очиқ мулоқотда бўлмоқда ва уларнинг муаммоларига ўз вақтида жавоб қайтармоқда" Мазкур фикрга...:',
            '14. “Мамлакатимизда олиб борилаётган ислоҳотлар тўғри йўлда кетмоқда”, мазкур фикрга:',
            '26. Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?',
            '27. Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?',
            '30. Сизнингча, Шавкат Мирзиёев мамлакатдаги вазиятни яхши томонга ўзгартира оладими ёки йўқми?']
    region_col = 'Яшаш ҳудудингиз:'
    df_out = pd.DataFrame()
    merge_columns = {}
    dfs = []
    sheet_names = []
    for i, col in enumerate(cols):
        
        ctab = pd.crosstab(index = df[region_col], columns = df[col])
        ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col], normalize='index') * 100).round(decimals=1)
        total_by_region = ctab.sum(axis=1)
        total_all = total_by_region.sum()
        total_by_region.name='Жами'
        # ctab = pd.concat([ctab, total_by_answers], axis=0)
        total_by_region= pd.DataFrame(total_by_region)
        ctab = pd.concat([total_by_region['Жами'], ctab], axis=1)
        
        total_by_answers = ctab.sum(axis=0)
        total_by_answers = total_by_answers.reset_index().T
        total_by_answers.rename(dict(zip(total_by_answers.columns.values.tolist(), total_by_answers.iloc[0].values.tolist())), inplace=True, axis=1)
       
        total_by_answers = total_by_answers.tail(-1)
        total_by_answers.index = ['Жами:']
        ctab = pd.concat([ctab, total_by_answers], axis=0)
        
        
        ctab = pd.concat([ctab['Жами'], ctab_pct], axis=1)
        ctab.drop(index='Жами:', inplace=True)
        total_by_answers_pct = (total_by_answers.div(total_all)).astype(float).multiply(100).round(decimals=1)
        
        # total_by_answers_pct['Жами'] = None
        total_by_answers_pct.index=['Жами, (%):']
        ctab = pd.concat([ctab, total_by_answers_pct], axis=0)
        ctab = pd.concat([ctab, total_by_answers], axis=0)

        
 
        # ctab = (pd.crosstab(index = df[region_col], columns = df[col], normalize='index') * 100).round(decimals=1)
        ctab[col] = ctab.index 
        ctab = mc.MoveTo1(ctab, col)
        ctab = ctab.reset_index(drop=True).T.reset_index().T      
        ctab.reset_index(drop=True, inplace=True)
        
        # ctab.style.background_gradient(cmap='coolwarm').set_precision(2)
        # ctab.show()
        merge_columns[ctab.iloc[0][0]] = len(ctab.columns)

        dfs.append(ctab)
        
        sn = f'Sheet{i+1}'
        if i == 0:
            with pd.ExcelWriter(out_filename, mode='w', engine='openpyxl') as writer:
                ctab.to_excel(writer, index=False, sheet_name=sn)
        else:
            with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl', if_sheet_exists='error') as writer:
                ctab.to_excel(writer, index=False, sheet_name=sn)

        sheet_names.append(sn)

        workbook = openpyxl.load_workbook(filename=out_filename)
        worksheet = workbook[sn]
        cell = worksheet.cell(row=1, column=1)
        cell.value = col
        cell = worksheet.cell(row=2, column=1)
        cell.value = 'Ҳудуд'
        workbook.save(out_filename)




    

    print(f'{datetime.today()} - Starting generating REGIONAL REPORT file...')

