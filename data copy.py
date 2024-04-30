

from datetime import datetime
from download import excel_to_pandas
from default import URL
from web import to_pdf
import pandas as pd 
from default import REPORT_COLS, outpath
import os 
import movecolumn as mc 
import numpy as np
import string
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import re
from utils.utils import get_rich_text, SIZE, FONT_NAME, HEADER_SIZE, crosstab
import warnings
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

def convert_to_dict_of_lists(lst):
    result_dict = {}
    for item in lst:
        key = item[0]
        value = item[1]
        if key in result_dict:
            result_dict[key].append(value)
        else:
            result_dict[key] = [value]
    return result_dict

root = os.getcwd()
report_cols_list = [REPORT_COLS[key] for key in list(REPORT_COLS.keys())]
alphabet = dict.fromkeys(string.ascii_uppercase, 0)
alphabet = dict(zip(range(len(list(alphabet.keys()))), list(alphabet.keys())))


def to_excel(df, name):
    name = os.path.join(outpath, name)
    df.to_excel(name, index=False)







def generate_pdf():
    today = str(datetime.today().date()).replace('-', '_')
    db_filename = 'db_{today}.xlsx'.format(today=today)
    df = excel_to_pandas(URL, 'data/' + db_filename)
    return df

    

def report(df):
    mode ='w'
    prefix = ''
    today = str(datetime.today().date()).replace('-', '_')
    report_filename = 'report_analysis_{today}.xlsx'.format(today=today)
    out_filename = f'out/report/{report_filename}'
    # df=pd.read_excel(f'out\\db_{today}.xlsx')

    df['_submission_time'] = pd.to_datetime(df['_submission_time']).dt.date
    
    mask_eff = ~pd.isnull(df['2.7. Ер ажратилган пайтда унинг унумдорлиги қандай эди?'])
    df = df[mask_eff]
    df.drop_duplicates(subset=['name', '1.1.1. Ҳудуд:', '1.1.2. Туман:', '1.1.3. Маҳалла:'], keep='last', inplace=True)

    df_orig = df.copy()
    df.rename(columns={'1.1.1. Ҳудуд:': 'region', '1.1.2. Туман:': 'district', '_id': 'count'}, inplace=True)
    df = df[['region', 'district', 'count', '_submission_time']]

    
    dates = df_orig['_submission_time'].unique()
    df_out = pd.DataFrame()
    for i, date in enumerate(sorted(dates)):
        mask = df['_submission_time'] == date
        df_temp = df[mask]
        df_temp.drop('_submission_time', axis=1, inplace=True)
        df_temp = df_temp.groupby(by=['district'], as_index=False).count()
        df_temp.rename(columns={'region':'Ҳудуд',  'district': 'Туман', 'count': date}, inplace=True)
        total = df_temp[date].sum()
        df_temp.drop('Ҳудуд', axis=1, inplace=True)
        
        
        if i == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='outer', on=['Туман'])
    
    
    total_col = df_out[dates.tolist()].sum(axis=1)
    df_out['Жами']= total_col

    # sys.exit()



    # df_total = pd.DataFrame(columns=df_out.columns, data=[['Жами:', np.nan, np.nan, np.nan, np.nan, np.nan, np.nan , total]])
    df_total = df_out[dates].sum()
    total_all = sum(df_total.values.tolist())
    df_total = pd.DataFrame(columns=['Туман'] +df_total.index.values.tolist(), data=[['Жами:']+df_total[dates].values.tolist()])
    
    df_out = pd.concat([df_out, df_total], axis=0)
    df_out.reset_index(drop=True, inplace=True)
    df_out.at[df_out.index.values.tolist()[-1], 'Жами'] = total_all
    df_out['Жами'] = df_out['Жами'].astype(int)
    df_plan = pd.read_excel('plan.xlsx', sheet_name='mahalla')[['mahalla', 'plan','region', 'district']]
    
    df_plan = df_plan.groupby(by=['region', 'district'], as_index=False)['plan'].sum()

    df_out = pd.merge(df_out, df_plan, left_on='Туман', right_on='district', how='right')
    df_out['Туман'] = df_out['district']
    df_out.drop(columns=['district'], inplace=True)
    
    df_out.rename(columns={'plan': 'Режа', 'region': 'Ҳудуд'}, inplace=True)
    df_out.set_index(['Ҳудуд', 'Туман'], inplace=True)
    # df_out.at[df_out.index.values[-1], 'Режа'] = df_out['Режа'].sum()
    df_out['Режа'] = df_out['Режа'].astype('Int64')

    df_out['Ҳолати, %'] = (df_out['Жами'].div(df_out['Режа'])).round(3) * 100
    df_out = df_out.sort_index()
    totals = df_out.sum()
    totals['Ҳолати, %'] = round(totals['Жами']/totals['Режа'], 3) * 100
    totals = totals.to_frame().T
    arrays = [["Жами"], ['']]

    totals.index = pd.MultiIndex.from_arrays(arrays, names=('Ҳудуд', 'Туман'))
    df_out = df_out.fillna(0)
    df_out = pd.concat([totals, df_out], axis=0)

    with pd.ExcelWriter(out_filename, mode=mode, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}districts')

    # print(df_out)
    ############################################################################ Вилоятлар    
    df_orig.rename(columns={'1.1.1. Ҳудуд:': 'region', '1.1.2. Туман:': 'district', '_id': 'count', '1.1.3. Маҳалла:': 'mahalla'}, inplace=True)
    df_orig = df_orig[['region', 'district', 'count', '_submission_time', 'mahalla']]

    # df_orig = df_orig.groupby(by=['region'], as_index=False).count()[['region', 'count']]
    # df_orig.rename(columns={'region':'Ҳудуд', 'count': 'Уй хўжаликлар сони'}, inplace=True)
    df_out = pd.DataFrame()
   
    for k, date in enumerate(dates):
        mask = df_orig['_submission_time'] == date
        df_temp = df_orig[mask]
        df_temp = df_temp.groupby(by=['region'], as_index=False).count()[['region', 'count']]
        total = df_temp['count'].sum()
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        df_total = pd.DataFrame(columns=df_temp.columns, data=[['Жами:', total]])
        
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        
        if k == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='outer', on=['Ҳудуд'])
    df_out['Жами'] = df_out[dates].sum(axis=0)
    total_row = df_out.sum()[dates]
    total_row["Ҳудуд"] = 'Жами:'
    total_row = total_row.reset_index().T.reset_index(drop=True)
    total_row.rename(dict(zip(total_row.columns.values.tolist(), total_row.iloc[0].values.tolist())), inplace=True, axis=1)
    total_row.drop(0, axis=0, inplace=True)
    df_out = pd.concat([df_out, total_row], axis=0)
    columns = ["Ҳудуд"] + sorted(dates.tolist()) + ['Жами']
    df_out = df_out[columns]
    df_out['Жами'] = df_out[dates].sum(axis=1)

    df_plan = pd.read_excel('plan.xlsx')
    df_out = pd.merge(df_out, df_plan, on='Ҳудуд', how='left')
    df_out['Ҳолати, %'] = (df_out['Жами'].div(df_out['Режа'])*100).round(1)
    with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}regions', index=False)
    print(f'{datetime.today()} - Starting generating REPORT file...')


    # df_orig.rename(columns={'region': 'Ҳудуд', 'district': 'Туман', 'mahalla': 'Маҳалла'}, inplace=True)
    df_plan = pd.read_excel('plan.xlsx', sheet_name='mahalla')[['mahalla', 'plan','region', 'district']]
    df_orig = pd.merge(df_orig, df_plan, on=['region', 'district', 'mahalla'], how='left')
    df_plan = df_plan.set_index(['region', 'district', 'mahalla'])
    ctab_mahalla = pd.crosstab(columns=df_orig['_submission_time'], index=[df_orig['region'], df_orig['district'],df_orig['mahalla']], margins=True, margins_name='Жами')
    ctab_mahalla.reset_index(inplace=True)
    ctab_mahalla = pd.merge(ctab_mahalla, df_plan, on=['region', 'district', 'mahalla'], how='right')
    ctab_mahalla['Ҳолати, %'] = ((ctab_mahalla['Жами'].div(ctab_mahalla['plan']))).round(3) * 100
    ctab_mahalla.rename(columns={'region': 'Ҳудуд', 'district': 'Туман', 'mahalla': 'Маҳалла', 'plan': 'Режа'}, inplace=True)
    ctab_mahalla.set_index(['Ҳудуд', 'Туман', 'Маҳалла'], inplace=True)
    ctab_mahalla = ctab_mahalla.sort_index()
    ctab_mahalla = ctab_mahalla.fillna(0)

    totals = ctab_mahalla.sum()
    totals = totals.to_frame().T
    arrays = [["Жами"], [''], ['']]
    totals.index = pd.MultiIndex.from_arrays(arrays, names=('Ҳудуд', 'Туман', 'Маҳалла'))
    totals['Ҳолати, %'] = round(totals['Жами']/totals['Режа'], 3) * 100
    ctab_mahalla = pd.concat([totals, ctab_mahalla], axis=0)
    


    # ds_total = pd.Series(dtype=int)

    with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
        ctab_mahalla.to_excel(writer, sheet_name=f'{prefix}mahalla')


    workbook = openpyxl.load_workbook(filename=out_filename)
    alignment = Alignment(horizontal='left', 
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True,
                indent=0) 
    for sn in workbook.sheetnames:
        worksheet = workbook[sn]
        if sn == 'districts':
            columns = ['A', 'B']
            
        elif sn == 'regions':
            columns = ['A']
        elif sn == 'mahalla':
            columns = ['A', 'B', 'C']

        for i, c in enumerate(columns):
            worksheet.column_dimensions[c].width = 30
         
    workbook.save(filename=out_filename)
    return out_filename







def report_all(df, *args, **kwargs):
    mode ='w'
    prefix = ''
    today = str(datetime.today().date()).replace('-', '_')
    report_filename = 'all_report_analysis_{today}.xlsx'.format(today=today)
    out_filename = f'out/report/{report_filename}'
    # df=pd.read_excel(f'out\\db_{today}.xlsx')
    # df.drop_duplicates(subset=['name'], keep='last', inplace=True)
    df.drop_duplicates(subset=['name', '1.1.1. Ҳудуд:', '1.1.2. Туман:', '1.1.3. Маҳалла:'], keep='last', inplace=True)

    df['_submission_time'] = pd.to_datetime(df['_submission_time']).dt.date
    df_orig = df.copy()
    df.rename(columns={'1.1.1. Ҳудуд:': 'region', '1.1.2. Туман:': 'district', '_id': 'count'}, inplace=True)
    df = df[['region', 'district', 'count', '_submission_time']]

    
    dates = df_orig['_submission_time'].unique()
    df_out = pd.DataFrame()
    for i, date in enumerate(sorted(dates)):
        mask = df['_submission_time'] == date
        df_temp = df[mask]
        df_temp.drop('_submission_time', axis=1, inplace=True)
        df_temp = df_temp.groupby(by=['district'], as_index=False).count()
        df_temp.rename(columns={'region':'Ҳудуд',  'district': 'Туман', 'count': date}, inplace=True)
        total = df_temp[date].sum()
        df_temp.drop('Ҳудуд', axis=1, inplace=True)
        
        
        if i == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='outer', on=['Туман'])
    
    
    total_col = df_out[dates.tolist()].sum(axis=1)
    df_out['Жами']= total_col

    # sys.exit()



    # df_total = pd.DataFrame(columns=df_out.columns, data=[['Жами:', np.nan, np.nan, np.nan, np.nan, np.nan, np.nan , total]])
    df_total = df_out[dates].sum()
    total_all = sum(df_total.values.tolist())
    df_total = pd.DataFrame(columns=['Туман'] +df_total.index.values.tolist(), data=[['Жами:']+df_total[dates].values.tolist()])
    
    df_out = pd.concat([df_out, df_total], axis=0)
    df_out.reset_index(drop=True, inplace=True)
    df_out.at[df_out.index.values.tolist()[-1], 'Жами'] = total_all
    df_out['Жами'] = df_out['Жами'].astype(int)
    df_plan = pd.read_excel('plan.xlsx', sheet_name='mahalla')[['mahalla', 'plan','region', 'district']]
    
    df_plan = df_plan.groupby(by=['region', 'district'], as_index=False)['plan'].sum()

    df_out = pd.merge(df_out, df_plan, left_on='Туман', right_on='district', how='right')
    df_out['Туман'] = df_out['district']
    df_out.drop(columns=['district'], inplace=True)
    
    df_out.rename(columns={'plan': 'Режа', 'region': 'Ҳудуд'}, inplace=True)
    df_out.set_index(['Ҳудуд', 'Туман'], inplace=True)
    # df_out.at[df_out.index.values[-1], 'Режа'] = df_out['Режа'].sum()
    df_out['Режа'] = df_out['Режа'].astype('Int64')

    df_out['Ҳолати, %'] = (df_out['Жами'].div(df_out['Режа'])).round(3) * 100
    df_out = df_out.sort_index()
    totals = df_out.sum()
    totals['Ҳолати, %'] = round(totals['Жами']/totals['Режа'], 3) * 100
    totals = totals.to_frame().T
    arrays = [["Жами"], ['']]

    totals.index = pd.MultiIndex.from_arrays(arrays, names=('Ҳудуд', 'Туман'))
    df_out = df_out.fillna(0)
    df_out = pd.concat([totals, df_out], axis=0)

    with pd.ExcelWriter(out_filename, mode=mode, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}districts')

    # print(df_out)
    ############################################################################ Вилоятлар    
    df_orig.rename(columns={'1.1.1. Ҳудуд:': 'region', '1.1.2. Туман:': 'district', '_id': 'count', '1.1.3. Маҳалла:': 'mahalla'}, inplace=True)
    df_orig = df_orig[['region', 'district', 'count', '_submission_time', 'mahalla', 'name']]

    # df_orig = df_orig.groupby(by=['region'], as_index=False).count()[['region', 'count']]
    # df_orig.rename(columns={'region':'Ҳудуд', 'count': 'Уй хўжаликлар сони'}, inplace=True)
    df_out = pd.DataFrame()
   
    for k, date in enumerate(dates):
        mask = df_orig['_submission_time'] == date
        df_temp = df_orig[mask]
        df_temp = df_temp.groupby(by=['region'], as_index=False).count()[['region', 'count']]
        total = df_temp['count'].sum()
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        df_total = pd.DataFrame(columns=df_temp.columns, data=[['Жами:', total]])
        
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        
        if k == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='outer', on=['Ҳудуд'])
    df_out['Жами'] = df_out[dates].sum(axis=0)
    total_row = df_out.sum()[dates]
    total_row["Ҳудуд"] = 'Жами:'
    total_row = total_row.reset_index().T.reset_index(drop=True)
    total_row.rename(dict(zip(total_row.columns.values.tolist(), total_row.iloc[0].values.tolist())), inplace=True, axis=1)
    total_row.drop(0, axis=0, inplace=True)
    df_out = pd.concat([df_out, total_row], axis=0)
    columns = ["Ҳудуд"] + sorted(dates.tolist()) + ['Жами']
    df_out = df_out[columns]
    df_out['Жами'] = df_out[dates].sum(axis=1)

    df_plan = pd.read_excel('plan.xlsx')
    df_out = pd.merge(df_out, df_plan, on='Ҳудуд', how='left')
    df_out['Ҳолати, %'] = (df_out['Жами'].div(df_out['Режа'])*100).round(1)
    with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}regions', index=False)
    print(f'{datetime.today()} - Starting generating REPORT file...')


    # df_orig.rename(columns={'region': 'Ҳудуд', 'district': 'Туман', 'mahalla': 'Маҳалла'}, inplace=True)
    df_plan = pd.read_excel('plan.xlsx', sheet_name='mahalla')[['mahalla', 'plan','region', 'district']]
    df_orig = pd.merge(df_orig, df_plan, on=['region', 'district', 'mahalla'], how='left')
    df_plan = df_plan.set_index(['region', 'district', 'mahalla'])
    ctab_mahalla = pd.crosstab(columns=df_orig['_submission_time'], index=[df_orig['region'], df_orig['district'],df_orig['mahalla']], margins=True, margins_name='Жами')
    ctab_mahalla.reset_index(inplace=True)
    ctab_mahalla = pd.merge(ctab_mahalla, df_plan, on=['region', 'district', 'mahalla'], how='right')
    ctab_mahalla['Ҳолати, %'] = ((ctab_mahalla['Жами'].div(ctab_mahalla['plan']))).round(3) * 100
    ctab_mahalla.rename(columns={'region': 'Ҳудуд', 'district': 'Туман', 'mahalla': 'Маҳалла', 'plan': 'Режа'}, inplace=True)
    ctab_mahalla.set_index(['Ҳудуд', 'Туман', 'Маҳалла'], inplace=True)
    ctab_mahalla = ctab_mahalla.sort_index()
    ctab_mahalla = ctab_mahalla.fillna(0)

    totals = ctab_mahalla.sum()
    totals = totals.to_frame().T
    arrays = [["Жами"], [''], ['']]
    totals.index = pd.MultiIndex.from_arrays(arrays, names=('Ҳудуд', 'Туман', 'Маҳалла'))
    totals['Ҳолати, %'] = round(totals['Жами']/totals['Режа'], 3) * 100
    ctab_mahalla = pd.concat([totals, ctab_mahalla], axis=0)

    ctab_mahalla = ctab_mahalla.reset_index()

    df_orig['address'] = df_orig['region'] + df_orig['district'] + df_orig['mahalla']
    address_list = df_orig['address'].unique().tolist()
    

    mask = (ctab_mahalla['Ҳудуд'] != 'Жами') & (ctab_mahalla['Ҳолати, %'] < 100)
    ctab_mahalla = ctab_mahalla[mask]

    


    ctab_mahalla['address'] = ctab_mahalla['Ҳудуд'] + ctab_mahalla['Туман'] + ctab_mahalla['Маҳалла']
    ctab_mahalla['done'] = False


    df_temp = pd.merge(df_orig, ctab_mahalla[['address', 'done']], how='left', on='address')

    df_temp = df_temp[df_temp['done'] == False]
    df_temp.drop_duplicates(subset=['address'], inplace=True)
    
    df_list = pd.read_csv('data\\hhidfile.csv')
    df_list['address'] = df_list['region']+df_list['district']+df_list['mahalla']
    

    df_add_empty = pd.Series()
    for i, row in df_list.drop_duplicates(subset=['address']).iterrows():
        if not row['address'] in address_list:
            df_add_empty['region'] = row['region']
            df_add_empty['district'] = row['district']
            df_add_empty['count'] = i + 999
            df_add_empty['_submission_time'] = None
            df_add_empty['name'] = 9999
            df_add_empty['plan'] = df_list[(df_list['mahalla'] == row['mahalla']) & (df_list['list_name'] == 'er_egasi')].shape[0]
            df_add_empty['address'] = row['address']
            df_add_empty['done'] = False
            df_temp = pd.concat([df_temp, df_add_empty.to_frame().T], axis=0)
    df_list = pd.merge(df_list, df_temp[['address', 'done']], how='left', on='address', suffixes=('_list', ''))
    df_list = df_list[~pd.isnull(df_list['done'])]
    for i, row in df_list.iterrows():
        if row['name'] in df_orig['name'].values.tolist() and row['address'] in df_orig['address'].values.tolist():
            df_list.at[i, 'done'] = True

    df_list = df_list[(df_list['done'] == False)]
    list_name = df_list['list_name']
    label = df_list['label']
    df_list.drop(columns=['name', 'pinfl', 'region_code', 'district_code', 'filter_province', 'filter_zaxira', 'address','done', 'mahalla_slug', 'list_name', 'label'], inplace=True)
    df_list['list_name'] = list_name
    df_list['label'] = label


    df_list_mahalla = df_list.drop_duplicates(subset=['region', 'district', 'mahalla'])
    df_list_mahalla.drop(columns=['list_name', 'label'], inplace=True)
    
    df_list['list_name'] = df_list['list_name'].replace('er_egasi', 'Асосий').replace('zaxira', 'Захира')

    df_list.rename(columns={'region': 'Ҳудуд', 'district':"Туман", 'mahalla':"Маҳалла", "list_name":"Рўйхат", 'label':"Исми"}, inplace=True)
    df_list_mahalla.rename(columns={'region': 'Ҳудуд', 'district':"Туман", 'mahalla':"Маҳалла"}, inplace=True)


    df_list.set_index(['Ҳудуд', 'Туман', 'Маҳалла', 'Рўйхат'], inplace=True)
    df_list.sort_index(inplace=True)

    df_report = pd.read_excel(kwargs['out_filename'], sheet_name='mahalla')
    df_report['Ҳудуд'] = df_report['Ҳудуд'].ffill()
    df_report['Туман'] = df_report['Туман'].ffill()
    df_list_mahalla = pd.merge(df_list_mahalla, df_report[['Ҳудуд', "Туман", "Маҳалла", "Жами", "Режа", "Ҳолати, %"]], how='left', on=['Ҳудуд', "Туман", "Маҳалла"])
    df_list_mahalla.set_index(['Ҳудуд', 'Туман', 'Маҳалла'], inplace=True)
    df_list_mahalla.sort_index(inplace=True)
    # ds_total = pd.Series(dtype=int)
    
    filename_nd = f'out\\report\\not_done_{today}.xlsx'
    with pd.ExcelWriter(filename_nd, mode='w', engine='openpyxl') as writer:
        df_list_mahalla.to_excel(writer, sheet_name=f'mahalla')
    
    with pd.ExcelWriter(filename_nd, mode='a', engine='openpyxl') as writer:
        df_list.to_excel(writer, sheet_name=f'list')

    with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
        ctab_mahalla.to_excel(writer, sheet_name=f'{prefix}mahalla')


    workbook = openpyxl.load_workbook(filename=out_filename)
    alignment = Alignment(horizontal='left', 
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True,
                indent=0) 
    for sn in workbook.sheetnames:
        worksheet = workbook[sn]
        if sn == 'districts':
            columns = ['A', 'B']
            
        elif sn == 'regions':
            columns = ['A']
        elif sn == 'mahalla':
            columns = ['A', 'B', 'C']

        for i, c in enumerate(columns):
            worksheet.column_dimensions[c].width = 30
         
    workbook.save(filename=out_filename)




    workbook = openpyxl.load_workbook(filename=filename_nd)
    alignment = Alignment(horizontal='left', 
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True,
                indent=0) 
    for sn in workbook.sheetnames:
        worksheet = workbook[sn]
        columns = ['A', 'B', 'C']

        for i, c in enumerate(columns):
            worksheet.column_dimensions[c].width = 30
         
    workbook.save(filename=filename_nd)
    return report_filename












def highlight(x):
    print(x)
    return x

def freq_table(df):
    drop_cols = [
        'deviceid',
        'audit',
        'audit_URL',
        '1.0. Локациянгизни жўнатинг:',
        '_1.0. Локациянгизни жўнатинг:_latitude',
        '_1.0. Локациянгизни жўнатинг:_longitude',
        '_1.0. Локациянгизни жўнатинг:_altitude',
        '_1.0. Локациянгизни жўнатинг:_precision',
        '_1.0. Локациянгизни жўнатинг:_latitude',
        '_id',
        '_uuid',
        '_submission_time',
        '_validation_status',
        '_notes',
        '_status',
        '_submitted_by',
        '__version__',
        '_tags',
        '_index',
        '2.2. Ушбу ерни қачон ижарага олгансиз?',
        '1.1.5. Ер эгасини танланг:',
        '1.1.3. Маҳалла:',
        '1.1.2. Туман:',
        '1.1.4. Респондентлар рўйхатини танланг:',
        '1.1.5. Ер эгасини танланг:',
        '1.8. Уй хўжалиги (хонадон) аъзолари сони:',
        '1.8.1. Хонадон аъзолари сонидан ИШСИЗЛАР (айни вақтда доимий даромадга эга бўлмаган, 18 ёшдан катта, ўқиш билан банд бўлмаган, пенсия ва нафақада бўлмаганлар):',
        'pivot_area',
        '5.6. Респондентнинг телефон рақами:',
        'pivot_area2',
        'name',
        'label'

    ]



    columns = df.columns.values.tolist()

    mult_choice = list(set([c.split(' ')[0] for c in columns if '/' in c]))
    mult_choice = [c for c in mult_choice if c.replace('.', '').strip(' ').isnumeric()]
    for c in mult_choice:
        for col in columns:
            if col.startswith(c) :
                drop_cols.append(col)
                break
                
    df.drop(columns=drop_cols, inplace=True)


    other_cols = [c for c in df.columns if 'Бошқа' in c and '**' in c]
    df.drop(columns=other_cols, inplace=True)
    
    
    df.rename(columns=dict(zip(df.columns.values.tolist(), columns)), inplace=True)

    # if 'out' not in df_filename:
    #     df_filename = os.path.join(root, 'out', df_filename)
    # df = pd.read_excel(df_filename)
    mask_eff = ~pd.isnull(df['2.4. Фойдаланаётган ер майдонини олишда “Е-аукцион”да бевосита ўзингиз ёки оила аъзоларингиз қатнашганми?'])
    df = df[mask_eff]
    df.dropna(axis=1, inplace=True, how='all')
    df_orig = df.copy()
    
    
    today = str(datetime.today().date()).replace('-', '_')
    
    
 
    df.reset_index(drop=True, inplace=True)

    columns = df.columns.values.tolist()
    df_out = pd.DataFrame()

    filename_out = f'out/freq/land_freq_table_{today}.xlsx'
    df_big = pd.DataFrame()

    town_vil = ['Аёл', 'Эркак','all',]
    
    # columns.pop(columns.index('_submission_time'))
    for i, c in enumerate(list(reversed(columns))):
        
        df_temp = pd.DataFrame()
        for j, tv in enumerate(town_vil):
            if tv in ['Шаҳар', 'Қишлоқ']:
                mask = df_orig['1. САВОЛ БЕРМАЙ, Қаерда (Шаҳар ёки қишлоқда) яшашини киритинг?'] == tv
                df = df_orig[mask]
            elif tv in ['18-35', '36-49', '50-999']:
                btm_age, top_age = tv.split('-')
                btm_age, top_age = int(btm_age), int(top_age)
                if btm_age == 18:
                    btm_age = 17
                btm_mask = df_orig['6. Ёшингиз?'] >= btm_age
                top_mask = df_orig['6. Ёшингиз?'] <= top_age
                # mask = ~(btm_mask & top_mask)
                df = df_orig[btm_mask][top_mask]
                df = df.dropna(subset='6. Ёшингиз?', axis=0)
                
            elif tv in ['Аёл', 'Эркак']:
                mask = df_orig['1.6. Респондентнинг жинси:'] == tv
                df = df_orig[mask]
            else:
                df = df_orig.copy()
            

            ctab = pd.crosstab(index = df.index, columns = df[c])
            ctab = ctab.sum()
            

            ctab_pct = (ctab/ctab.values.sum()).round(decimals=4)
            
            tab = pd.concat([ctab, ctab_pct], axis=1)
            
            cols = ['Frequency', 'Percentage (%)', c]
            tab[c] = tab.index
            
            
            df_out = pd.DataFrame(columns=cols, data=tab.values)
            df_out = mc.MoveTo1(df_out, c)
            

            if not df[c].dtype == 'object':
                df_out['Average score'] = df_out['Percentage (%)'] * df_out[c]
                try:
                    df_out['Average score'] = df_out['Average score'].round(decimals=3)
                except:
                    df_out['Average score'] = [np.nan] * len(df_out['Percentage (%)'])
            else:
                df_out['Average score'] = [np.nan] * len(df_out['Percentage (%)'])
            
            
            fillna_cols = df_out.columns.values.tolist()
            
            df_out[['Percentage (%)']] = df_out[['Percentage (%)']] * 100
            

            
            freq_sum = df_out['Frequency'].sum()
            pct_sum =df_out['Percentage (%)'].sum()
            avg_sum = df_out['Average score'].sum()
            if avg_sum == 0:
                avg_sum = np.nan
            df_out['Frequency'] = df_out['Frequency'].astype(float).astype('Int64')
            df_total = pd.DataFrame(columns=df_out.columns, data=[['Total:', freq_sum,\
                                                                pct_sum,
                                                                 avg_sum]])
            
            

            df_out = pd.concat([df_out, df_total], axis=0)
       
            

            
            # Add space after tables
            df_out.rename(columns={'Frequency': f'{tv}_Frequency', 'Percentage (%)': f'{tv}_Percentage (%)', 'Average score': f'{tv}_Average score'}, inplace=True)
            if j == 0:
                df_temp = df_out
            else:
                if tv == 'all' or tv in ['Аёл'] or tv == '18-35':
                    space = pd.DataFrame(np.nan, index=list(range(len(df_out))), columns=['A', 'B'])
                    # space = df_out[df_out.columns.values[:2]].fill(0)
                    df_out.reset_index(drop=True, inplace=True)
                    df_out = pd.concat([space, df_out], axis=1)
                
         
                df_temp = pd.merge(df_temp, df_out, how='outer', on=c)
                fillna_cols = df_temp.columns.values.tolist()
                fillna_cols = [fc for fc in fillna_cols if fc != 'Average score' and fc != c and fc not in ['A', 'B'] and 'A_' not in fc and 'B_' not in fc]
                df_temp[fillna_cols] = df_temp[fillna_cols].fillna(0)
            

        df_space = pd.DataFrame(columns=df_temp.columns.values.tolist(), data=[[np.nan]*len(df_temp.columns),
                                                                                [np.nan]*len(df_temp.columns)])
        
        ### move the Total row to the last
        col_first = df_temp.columns.values[0]
        mask = df_temp[col_first] == 'Total:'
        df_temp = pd.concat([df_temp[~mask], df_temp[mask]], axis=0)

        ## fill nan values with zeroes
        # df_temp.fillna(0, inplace=True)

        ## add space between question blocks
        df_temp = pd.concat([df_temp, df_space], axis=0)

        

        # columns = [c] + ['Frequency_x', 'Frequency_y', 'Percentage (%)_x',  'Percentage (%)_y', 'Average score_x', 'Average score_y', 'A', 'B', 'Frequency', 'Percentage (%)', 'Average score']
        # print(df_temp.columns.values)
        # sys.exit()
        columns = [c] + [ 'Аёл_Frequency', 'Эркак_Frequency', 'Аёл_Percentage (%)', 'Эркак_Percentage (%)', 'Аёл_Average score', 'Эркак_Average score', 'A', 'B',
                          'all_Frequency', 'all_Percentage (%)', 'all_Average score']
        


        

        micolumns = pd.MultiIndex.from_tuples(
                                            [
                                             (None, c),
                                             ("Frequency", "Аёл"), (None, "Эркак"), 
                                             ("Percentage (%)", "Аёл"), (None, "Эркак"),
                                             ("Average score", "Аёл"), (None, "Эркак"),
                                             (None, None),
                                             (None, None),
                                             ('Across all data', "Frequency"), 
                                             (None, "Percentage (%)"),
                                             (None, "Average score")],)
        
        # columns = [c] + ['all_Frequency', 'all_Percentage (%)', 'all_Average score']
   
        # micolumns = pd.MultiIndex.from_tuples(
        #                                     [(None, c),
        #                                      ('Across all data', "Frequency"), 
        #                                      (None, "Percentage (%)"),
        #                                      (None, "Average score")
        #                                      ],)

        df_temp = df_temp[columns]
        df_temp = pd.DataFrame(columns=micolumns, data=df_temp.values)

    
        
        
 
        if i == 0:
            df_big = df_temp
        else:
            
            df_temp = df_temp.reset_index(drop=True).T.reset_index().T
            df_temp.rename(columns=dict(zip(df_temp.columns.values.tolist(), df_big.columns.values.tolist())), inplace=True)
            if i == 1:
                columns = df_big.columns.values.tolist()
                df_big = df_big.reset_index(drop=True).T.reset_index().T
                df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), columns)), inplace=True)
            
            df_big = pd.concat([df_temp, df_big], axis=0)

    
        

    new_columns = [np.nan] * len(df_big.columns.values)
    new_columns[0] = 'Land Survey Frequency table'
    new_columns[6] = today.replace('_', '-')
    df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), new_columns)), inplace=True)
    
    with pd.ExcelWriter(filename_out, mode='w', engine='openpyxl') as writer:
        df_big.to_excel(writer, index=False)



    print(f'{datetime.today()} - Starting generating FREQ TABLE file...')
    return filename_out
    

def report_by_region(df):
    # df.rename(columns={'33.  Сизнингча, (вилоят ҳокими) вилоят/республика ҳокими сифатида қандай ишлаяпти?': '33. Вилоят хокими қандай ишлаяпти?',
    #                    '36. Сизнингча, туманингиз ҳокими ўз лавозимида қандай ишламоқда?': '36. Туман ҳокими ўз лавозимида қандай ишламоқда?',
    #                    "40. Ўзбекистон Республикаси Олий Мажлис Қонунчилик палатасининг амалдаги таркибининг фаолиятини қандай баҳолайсиз?": "40. Олий Мажлис Қонунчилик палатасининг амалдаги таркибининг фаолиятини қандай баҳолайсиз?",
    #                    "7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?": "7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангиз?",
    #                    "8. Вилоятингиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?": "8. Вилоятингиздаги умумий вазиятдан қониқиш даражангиз?",
    #                    "10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?": "10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангиз?",
    #                    "7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?": "7.2. Мамлакатдаги умумий вазият",
    #                    "9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?":"9. Вилоятдаги умумий вазият",
    #                    "11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?":"11. Маҳалладаги умумий вазият",}, inplace=True)

    today = str(datetime.today().date()).replace('-', '_')
    out_filename = f'out/regional/regional_{today}.xlsx'
    
    # cols = ['7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?',
    #         '7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?',
    #         '11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?',
    #         '9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?',
    #         '12. Сизнингча, атрофингиздаги одамлар ҳозир қандай кайфиятда: кўтаринки, хотиржам ёки тушкун (хавотирли)?',
    #         '13. “Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:',
    #         '16. “Ҳукумат фуқаролар билан очиқ мулоқотда бўлмоқда ва уларнинг муаммоларига ўз вақтида жавоб қайтармоқда" Мазкур фикрга...:',
    #         '14. “Мамлакатимизда олиб борилаётган ислоҳотлар тўғри йўлда кетмоқда”, мазкур фикрга:',
    #         '26. Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?',
    #         '27. Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?',
    #         '30. Сизнингча, Шавкат Мирзиёев мамлакатдаги вазиятни яхши томонга ўзгартира оладими ёки йўқми?']

    cols = {
        'single': ['1.1.6. (БУНИ ЎҚИМАНГ) Респондент сўровномада қатнашишга розими?',
                   '1.2. Респондентнинг ер эгасига алоқадорлиги:',
                   '1.3. 2022 йилда ушбу ердан ўзингиз фойдаландингизми?',
                   '2.3. Деҳқончилик қилиш, экинларни етиштириш бўйича тажрибангиз борми?',
                   '2.4. Фойдаланаётган ер майдонини олишда “Е-аукцион”да бевосита ўзингиз ёки оила аъзоларингиз қатнашганми?',
                   '2.7. Ер ажратилган пайтда унинг унумдорлиги қандай эди?',
                   '2.8. Ажратилган ерда сув таъминоти қандай?',
                   '4.5. Қўшимча молиявий ресурсларга (кредитга) эҳтиёжингиз борми ?',
                   '5.1. Ушбу ер майдонларига эгалик қилиш (фойдаланиш) уй хўжалигингизнинг даромадини .... :',
                   '5.3. Ушбу ердан фойдаланиш уй хўжалигингизнинг (хонадонингизнинг) озиқ-овқат таъминотига қандай таъсир қилди?',],
        'multiple': [
                     '1.9. (АГАР ҲОКИМ ЁРДАМЧИСИ БУНИ БИЛСА САВОЛНИ ЎҚИМАСДАН ЎЗИ ТЎЛДИРСИН) Сиз ёки уй хўжалигингиз аъзолари “ТЕМИР дафтар”, “АЁЛЛАР дафтари”, “ЁШЛАР дафтари”да рўйхатда турадими ёки илгари турганми?',
                     '2.5. Ер олишда қандай қийинчиликларга дуч келдингиз?',
                     '2.9. Ер майдонларини суғоришда сув манбасини кўрсатинг:',
                     '2.10. Суғоришнинг қайси усулидан фойдаланасиз?',
                     '2.11. Сув таъминоти билан боғлиқ қандай муаммо мавжуд?',
                     '3.3. Қайси турдаги қишлоқ хўжалиги экинларини экиш тўғрисида қандай қарор қилдингиз?',
                     '4.1. Ҳосилдорликни ошириш бўйича қандай ишларни амалга оширдингиз?',
                     '4.2. Фаолиятингизга Қишлоқ хўжалигида хизматлар кўрсатиш агентлиги (ёки Туман ҳокимлиги) қандай ёрдам берди?',
                     '4.3. Фаолиятингизда қандай муаммоларга дуч келмоқдасиз?',
                     '4.4. Сизнингча, ерлардан унумли фойдаланиш учун давлат томонидан яна қандай ишлар амалга оширилиши керак?',
                     '4.7. Қандай қонунбузилиш ҳолатларидан хабардорсиз?',
                     '5.4. 2022 йилда ердан олинган даромад қуйидаги қайси харажатларни ошириш имконини берди?',

                     ],
        'multicolumn': ['Биринчи экин',
                        'Иккинчи экин',
                        "Харажатлар",
                        "Кредитга бўлган этиёж ва даромад кўпайиши",
                        "Ер статистикаси"],
        

    }
    region_col = '1.1.1. Ҳудуд:'
    df_out = pd.DataFrame()
    merge_columns = {}
    dfs = []
    sheet_names = []
    counter = 0
    cells_ = {}


    for type_idx, (type_, columns_) in enumerate(cols.items()):
        for sheet_idx, col in enumerate(columns_):
            if type_ == 'single':
                ctab = crosstab(index=df[region_col], columns=df[col])

            elif type_ == 'multiple':
                if col == ' балл':
                    columns = [
                        '25. Президент Шавкат Мирзиёев ўз лавозимида фикрингиз бўйича қандай фаолият кўрсатаётганлигини 7 баллик шкалада баҳоланг?',
                        '33. Вилоят хокими қандай ишлаяпти?',
                        "36. Туман ҳокими ўз лавозимида қандай ишламоқда?",
                        "40. Олий Мажлис Қонунчилик палатасининг амалдаги таркибининг фаолиятини қандай баҳолайсиз?",
                        "7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангиз?",
                        "8. Вилоятингиздаги умумий вазиятдан қониқиш даражангиз?",
                        "10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангиз?"
                    ]
                                    
                    
                else:
                    columns = [c for c in df.columns.values if col in c][1:]
                    
                    drop_by_key = ['Ҳеч бири',
                                   "Бошқа",
                                   "Жавоб бериш",
                                   "Ҳеч қандай",
                                   "Телефон алоқаси билан боғлиқ муаммо",
                                   'Конвертация очилганлиги (валюта сиёсати)']
                    for key in drop_by_key:
                        columns = [c for c in columns if not key in c]
                    
                    if col == '28. Илтимос, Шавкат Мирзиёевнинг Президентлик даврида эришган асосий ютуқларини номлаб беринг?':
                        prepend_list = ['Янги Конституция қабул қилиниши',
                                 'Сўз эркинлигининг кучайиши',
                                 'Ёшларни қўллаб-қувватлашнинг кучайтирилиши',
                                 'Дин ва эътиқод эркинлигининг кучайиши',
                                 'Тадбиркорликни қўллаб-қувватлаш ишлари',
                                 'Камбағалликка қарши кураш ишлари',
                                 'Халқ билан мулоқотнинг кучайиши',
                                 'Соғлиқни сақлаш тизимидаги ислоҳотлар',
                                 ]
                                 
                        columns_prepend = []
                        
                        for p in prepend_list:
                            for i, c in enumerate(columns[:]):
                                if p in c:
                                    columns_prepend.append(c)
                                    columns.pop(i)
                                    break
                        columns = columns_prepend + columns
                        
                dfs_multiple = {}
                for col_ in columns: 
                    # ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col_], normalize='index', margins=True) * 100).round(decimals=1)
                    # index = ctab_pct.index.values.tolist()
                    # index[-1] = 'Республика бўйича'
                    # ctab_pct.index = index
                    # ctab_pct = pd.concat([ctab_pct.tail(1), ctab_pct.head(-1)], axis=0)
                    # ctab = ctab_pct
                    
                    # # total_by_answers_pct['Жами'] = None
                    col_name = col_.replace(col+'/', '')
                    
                    ctab = crosstab(index = df[region_col], columns = df[col_], rename={1:col_name}, remove_names=False)
                    if col == ' балл':
                        col_name = col_name.split(' ', maxsplit=1)[1]

                        column_values = pd.Series(ctab.columns.values.tolist(), index=ctab.columns)
                        ctab = ctab/100
                        ctab = ctab.multiply(column_values, axis=1)
                        ctab['Average score'] = ctab.sum(axis=1)
                        ctab = ctab['Average score'].to_frame().round(1)
                        ctab.rename(columns={'Average score': 1}, inplace=True)

                    if not (col_name,    '%') in ctab.columns:
                        ctab.columns = ctab.columns.set_levels([col_name, None], level=0)
                        ctab[(col_name,    '%')] = 0
                    ctab.rename(columns={1:col_name}, inplace=True)
                    
                    # dfs_multiple.append(ctab[col_name])#.to_frame())
                    
                    dfs_multiple[col_name] = ctab[col_name]#.to_frame())
                
                
                ctab = pd.concat(dfs_multiple.values(), axis=1, keys=dfs_multiple.keys())
            elif type_ == 'multicolumn':
                
                # columns = [c for c in df.columns.values if col in c]
                if col == 'Биринчи экин':
                    columns = ['3.3.1. Дон ва дуккакли экинлар (Буғдой, арпа, шоли, нўҳат, ловия) учун ажратилган ер майдонини киритинг:',
                                '3.3.1.1. Дон ва дуккакли экинлардан (Буғдой, арпа, шоли, нўҳат, ловия) олган ҳосил миқдорини киритинг:',
                                '3.3.1.2.  Дон ва дуккакли экинлардан (Буғдой, арпа, шоли, нўҳат, ловия) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.3.3. Сабзавотлар (помидор, бодринг, сабзи, пиёз ва бошқалар) учун ажратилган ер майдонини киритинг:',
                                '3.3.3.1. Сабзавотлардан (помидор, бодринг, сабзи, пиёз ва бошқалар) олган ҳосил миқдорини киритинг:',
                                '3.3.3.2. Сабзавотлардан (помидор, бодринг, сабзи, пиёз ва бошқалар) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.3.4. Полиз экинлари (тарвуз, қовун) учун ажратилган ер майдонини киритинг:',
                                '3.3.4.1. Полиз экинларидан (тарвуз, қовун) олган ҳосил миқдорини киритинг:',
                                '3.3.4.2. Полиз экинларидан (тарвуз, қовун) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.3.5. Картошка учун ажратилган ер майдонини киритинг:',
                                '3.3.5.1. Картошкадан олган ҳосил миқдорини киритинг:',
                                '3.3.5.2. Картошкадан олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.3.6. Озуқа экинлари учун ажратилган ер майдонини киритинг:',
                                '3.3.6.1. Озуқа экинларидан олган ҳосил миқдорини киритинг:',
                                '3.3.6.2. Озуқа экинларидан олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.3.7. Бошқа экинлар (тамаки, лавлаги, мойли экинлар) учун ажратилган ер майдонини киритинг:',
                                '3.3.7.1. Бошқа экинлардан (тамаки, лавлаги, мойли экинлар) олган ҳосил миқдорини киритинг:',
                                '3.3.7.2. Бошқа экинлардан (тамаки, лавлаги, мойли экинлар) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.4.1. Ерга ишлов бериш (ерни текислаш, хайдаш, экинга тайёрлаш) учун сарфлаган харажатингиз?',
                                '3.4.2. Уруғлик ва кўчат учун сарфлаган харажатингиз?',
                                '3.4.3. Ўғитлар учун сарфлаган харажатингиз?',
                                '3.4.4. Пестицидлар (зарар кунандаларга қарши курашиш воситалари) учун сарфлаган харажатингиз?',
                                '3.4.5. Суғориш учун сарфлаган харажатингиз?',
                                '3.4.6. Ишчи кучи (мавсумий ишчилар) учун сарфлаган харажатингиз?',
                                '3.4.7. Бошқа харажатлар:',
                                '3.5. 1-экиндан жами даромадингиз?',
                               ]
                if col == 'Иккинчи экин':
                    columns = ['3.6.1. Дон ва дуккакли экинлар (Буғдой, арпа, шоли, нўҳат, ловия) учун ажратилган ер майдонини киритинг:',
                                '3.6.1.1. Дон ва дуккакли экинлардан (Буғдой, арпа, шоли, нўҳат, ловия) олган ҳосил миқдорини киритинг:',
                                '3.6.1.2. Дон ва дуккакли экинлардан (Буғдой, арпа, шоли, нўҳат, ловия) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.6.3. Сабзавотлар (помидор, бодринг, сабзи, пиёз ва бошқалар) учун ажратилган ер майдонини киритинг:',
                                '3.6.3.1. Сабзавотлардан (помидор, бодринг, сабзи, пиёз ва бошқалар) олган ҳосил миқдорини киритинг:',
                                '3.6.3.2. Сабзавотлардан (помидор, бодринг, сабзи, пиёз ва бошқалар) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.6.4. Полиз экинлари (тарвуз, қовун) учун ажратилган ер майдонини киритинг:',
                                '3.6.4.1. Полиз экинларидан (тарвуз, қовун) олган ҳосил миқдорини киритинг:',
                                '3.6.4.2. Полиз экинларидан (тарвуз, қовун) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.6.5. Картошка учун ажратилган ер майдонини киритинг:',
                                '3.6.5.1. Картошкадан олган ҳосил миқдорини киритинг:',
                                '3.6.5.2. Картошкадан олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.6.6. Озуқа экинлари учун ажратилган ер майдонини киритинг:',
                                '3.6.6.1. Озуқа экинларидан олган ҳосил миқдорини киритинг:',
                                '3.6.6.2. Озуқа экинларидан олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.6.7. Бошқа экинлар (тамаки, лавлаги, мойли экинлар) учун ажратилган ер майдонини киритинг:',
                                '3.6.7.1. Бошқа экинлардан (тамаки, лавлаги, мойли экинлар) олган ҳосил миқдорини киритинг:',
                                '3.6.7.1. Бошқа экинлардан (тамаки, лавлаги, мойли экинлар) олинган ҳосилнинг қанча қисмини ўзингиз истеъмол қилдингиз?',
                                '3.7.1. Ерга ишлов бериш (ерни текислаш, хайдаш, экинга тайёрлаш) учун сарфлаган харажатингиз?',
                                '3.7.2. Уруғлик ва кўчат учун сарфлаган харажатингиз?',
                                '3.7.3. Ўғитлар учун сарфлаган харажатингиз?',
                                '3.7.4. Пестицидлар (зарар кунандаларга қарши курашиш воситалари) учун сарфлаган харажатингиз?',
                                '3.7.5. Суғориш учун сарфлаган харажатингиз?',
                                '3.7.6. Ишчи кучи (мавсумий ишчилар) учун сарфлаган харажатингиз?',
                                '3.7.7. Бошқа харажатлар:',
                                '3.8. 2-экиндан жами даромадингиз?',
                            ]
                if col == 'Харажатлар':
                    columns = [
                                '3.10. Ер учун бир йиллик ижара тўловингиз қанча?',
                                '3.11. Ер учун бир йиллик жами солиқ тўловингиз қанча?',
                                '3.12.1. Тупроқ сифатини яхшилаш учун сарфланган харажатлар миқдори?',
                                '3.12.2. Сув таъминотини яхшилаш учун сарфланган харажатлар миқдори?',
                                '3.12.3. Қишлоқ хўжалиги техникасини харид қилиш учун сарфланган харажатлар миқдори?',
                                '3.12.4. Бошқа (консалтинг хизматлари, ер атрофини ўраш в.б.) харажатлар миқдори?',
                                "2.12. Суғориш учун ЙИЛЛИК харажатингизни кўрсатинг:",]

                if col == 'Кредитга бўлган этиёж ва даромад кўпайиши':
                    columns = ['4.6. Қанча миқдорда қредитга эҳтиёжингиз бор?',
                                '5.2. Ушбу ердан фойдаланиш уй хўжалигининг (хонадоннинг) ўртача йиллик даромадини қанчага кўпайтирди?#####(олдинги ва ҳозирги даромаднинг фарқини кўрсатинг)',
                                ]
                if col == 'Ер статистикаси':
                    columns = ['1.4.  Ажратилган ер майдони қанча?',
                                '2.6. Ажратилган ер майдонлари сиз яшаб турган ҳудуддан неча км узоқликда жойлашган?',
                                "3.2. 2022 йилда неча марта маҳсулот етиштиргансиз?"]
                
                
                if '“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:' in columns:
                    col_idx = columns.index('“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:')
                    columns[col_idx] = '“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:'
                
                new_columns = [(c.split(' ', maxsplit=1)[1].replace('...', '')).strip() for c in columns]
                for c_idx, c in enumerate(new_columns):
                    if c == 'Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?':
                        new_columns[c_idx] = 'Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими?'
                    elif c == 'Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?':
                        new_columns[c_idx] = 'Шавкат Мирзиёевга ишонасизми?'
                        
                    
                    
                
                # df.rename(columns=dict(zip(columns, new_columns)), inplace=True)
                # columns = new_columns
                dfs_multiple = []
                data = {}

                for col_ in columns:
                    
                    # ctab = pd.crosstab(index = df[region_col], columns = df[col_])
                    # ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col_], normalize='index', margins=True) * 100).round(decimals=1)
                    # index = ctab_pct.index.values.tolist()
                    # index[-1] = 'Республика бўйича'
                    # ctab_pct.index = index
                    # ctab_pct = pd.concat([ctab_pct.tail(1), ctab_pct.head(-1)], axis=0)
                    # ctab = ctab_pct
                    # col_name = col_.replace(col+'/', '')
                    data = df.groupby(by=['1.1.1. Ҳудуд:'], as_index=True)[['1.1.1. Ҳудуд:', col_]].mean()
                    total_data = (df[col_].mean())
                    total_data= pd.DataFrame(columns=['1.1.1. Ҳудуд:', col_], data=[['Республика бўйича', total_data]])
                    total_data.set_index('1.1.1. Ҳудуд:', inplace=True)
                    data = pd.concat([total_data, data], axis=0)

                    if data.mean().values[0] < 200:
                        data = data.round(1)
                    else:
                        data = data.round(0)
                    
                        

                    # print(ctab)
                    # ctab = pd.DataFrame(data, index=ctab.index)
                    dfs_multiple.append(data)
                ctab = pd.concat(dfs_multiple, axis=1)
            

            # ctab = (pd.crosstab(index = df[region_col], columns = df[col], normalize='index') * 100).round(decimals=1)
            ctab[col] = ctab.index 
            ctab = mc.MoveTo1(ctab, col)
            ctab = ctab.reset_index(drop=True).T.reset_index().T      
            ctab.reset_index(drop=True, inplace=True)
            if type_ != 'multicolumn':
                for ci in ctab.columns:
                    if ci != 0 and ci%2 == 0:
                        if ctab.at[0, ci] == ctab.at[0, ci-1]:
                            ctab.at[0, ci] = None
            
            # ctab.style.background_gradient(cmap='coolwarm').set_precision(2)
            # ctab.show()
            merge_columns[ctab.iloc[0][0]] = len(ctab.columns)

            dfs.append(ctab)
            
            sn = col
            if len(col) > 31:
                sn = col[:-31]
            if counter == 0:
                with pd.ExcelWriter(out_filename, mode='w', engine='openpyxl') as writer:
                    ctab.T.reset_index().T.to_excel(writer, index=False,header=None, sheet_name=sn)
            else:
                with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl', if_sheet_exists='error') as writer:
                    ctab.T.reset_index().T.to_excel(writer, index=False,header=None, sheet_name=sn)
            
         
            counter +=1
            sheet_names.append(sn)
            
            # removing question nums from headers
            if col[0].isnumeric():
                col = col.split(' ', maxsplit=1)[1] 
            workbook = openpyxl.load_workbook(filename=out_filename)
            worksheet = workbook[sn]
            worksheet.sheet_view.zoomScale = 55
            ft_default = Font(size=SIZE, name=FONT_NAME)
            ft_blue = Font(color="0070c0", bold=True, size=SIZE, name=FONT_NAME)
            ft_red = Font(color="C00000", bold=True, size=SIZE, name=FONT_NAME)
            ft_column = Font(color='1f4e78', bold=True, size=SIZE, name=FONT_NAME)
            ft_header = Font(color='1f4e78', bold=True, size=HEADER_SIZE, name=FONT_NAME)
            ft_tiny = Font(color='1f4e78', italic=True, size=14, name=FONT_NAME)
            ft_bold = Font(color='000000', bold=True, italic=False, size=SIZE, name=FONT_NAME)
            edge_col = len(ctab.columns)
            for c in range(edge_col):
                cell = worksheet.cell(row=1, column=c+1)
                cell.value = None
            cell = worksheet.cell(row=1, column=1)
     


            cell.value = col
          
            
            cell.font = ft_header
            cell.alignment = Alignment(horizontal='center', 
                                vertical='center',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0) 
            worksheet.row_dimensions[2].height = 50
            worksheet.insert_rows(1)
            worksheet.insert_rows(3)
            worksheet.insert_rows(3)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=edge_col)

            
            none_columns = [(column_i, column_i-1) for column_i, column_ in enumerate(ctab.iloc[0].values.tolist()) if column_ is None]
            
            
            
            cell = worksheet.cell(row=5, column=1)
            cell.value = None
            # worksheet.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)

            worksheet.column_dimensions['A'].width = 34
            # cell = worksheet.cell(row=1, column=edge_col)
            # cell.value = f'{counter}-жадвал'

            cell = worksheet.cell(row=4, column=edge_col)
            if type_ == 'multicolumn':
                cell.value = f'(ўртача қиймат)'
            else:
                cell.value = f'(фоизда)'

            cell.alignment = Alignment(horizontal='right', 
                                vertical='bottom',
                                text_rotation=0,
                                wrap_text=False,
                                shrink_to_fit=False,
                                indent=0) 

            
            cell.font = ft_tiny

            if type_ == 'multicolumn':
                worksheet.row_dimensions[6].height = 60
                # columns = ctab.iloc[0].tolist()[1:]
                # all_cols = ctab.iloc[1].values.tolist()[1:]
                # columns = list(zip(columns, all_cols))
                # columns = convert_to_dict_of_lists(columns)
                key_count = 1
                # for _, col_ in enumerate(list(columns.keys())):
                #     target_cols = columns[col_]
                #     for ti, target_col in enumerate(target_cols):
                #         key_count += 1
                #         if ti == len(target_cols)-1:
                #             worksheet.merge_cells(start_row=5, start_column=key_count-len(target_cols)+1, end_row=5, end_column=key_count)
                #             worksheet[f'{get_column_letter(key_count-len(target_cols))}5'] = col_
                #             break
                        
                #         cell = worksheet.cell(row=5, column=key_count)
                #         cell.value = None
                        
                    
                    # worksheet[f'{alphabet[key_count-1]}5'] = col_

                    # cell = worksheet.cell(row=6, column=key_count)
                    # cell.font = ft_column

                worksheet.row_dimensions[6].height = 60
            
            for j in range(5, 22):
                alignment = Alignment(horizontal='general', 
                                vertical='bottom',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0) 
                cell = worksheet.cell(row=j, column=1)
                cell.alignment = alignment
                worksheet.row_dimensions[j].height = 30
                if j == 5:
                    worksheet.row_dimensions[j].height = 120
                    
                for k in range(1, edge_col+1):
                    cell = worksheet.cell(row=j, column=k)
                    cell.font = ft_default
                    if k == 1 and j == 6:
                        cell.font = ft_column
                        continue
                    if k != 1:
                        alignment.horizontal = 'center' 
                        alignment.vertical = 'center' 
                    if j == 5:
                        if len(ctab.columns) > 11:
                            if not col in ['Ер олишда қандай қийинчиликларга дуч келдингиз?', '2022 йилда ердан олинган даромад қуйидаги қайси харажатларни ошириш имконини берди?',\
                                           'Фаолиятингизда қандай муаммоларга дуч келмоқдасиз?', 'Сизнингча, ерлардан унумли фойдаланиш учун давлат томонидан яна қандай ишлар амалга оширилиши керак?']:
                                alignment.text_rotation = 90
                                worksheet.row_dimensions[j].height = 190
                            worksheet.column_dimensions[get_column_letter(k)].width = 13
                        else:
                            worksheet.column_dimensions[get_column_letter(k)].width = 20
                            

                        cell.font = ft_column
                    
                    
                    else:  
                        rep_cell = worksheet.cell(row=6, column=1)
                        rep_vals = []
                        rep_row = None
                        resp_mask = ctab[0] == 'Республика бўйича'
                        resp_idx = ctab[resp_mask].index.values[0]+1+4
                        resp_idx_ctab = ctab[resp_mask].index.values[0]
                        resp_values = []
                        for __c__ in ctab.columns:
                            if ctab.at[resp_idx_ctab-1, __c__] != 'Сони':
                                if not isinstance(ctab[resp_mask][__c__].values[0], str):
                                    resp_values.append(ctab[resp_mask][__c__].values[0])
                        

                        resp_mean = np.mean(resp_values)
                        thresh = resp_mean
                        lower_thresh = resp_mean
                        ft_high = ft_blue
                        ft_low = ft_red
                        
                        
                            
                        
                        if rep_cell.value == 'Республика бўйича':
                            rep_cell.font = ft_column
                            rep_vals = [ (worksheet.cell(row=6, column=c_).value, c_)  for c_ in range(1, edge_col+1) if worksheet.cell(row=5, column=c_).value != 'Сони'][1:]
                            rep_vals = [v[1] for v in rep_vals if v[0] >=thresh]
                            rep_row = 6
                        elif worksheet.cell(row=7, column=1).value == 'Республика бўйича':
                            rep_cell = worksheet.cell(row=7, column=1)
                            rep_cell.font = ft_column
                            rep_vals = [(worksheet.cell(row=7, column=c_).value, c_)  for c_ in range(1, edge_col+1) if worksheet.cell(row=6, column=c_).value != 'Сони'][1:]
                            
                            rep_vals = [v[1] for v in rep_vals if v[0] >=thresh]
                            rep_row = 7
                            
                        
                        
                        if not cell.value is None:
                            col_highlight = []
                            if not isinstance(cell.value, str):
                                if j  == rep_row:
                                    cell.font = ft_bold
                                if k in rep_vals:
                                    if cell.value >= thresh:
                                        cell.font = ft_high
                                    elif cell.value <=lower_thresh:
                                        cell.font = ft_low
                                                                            
                    
                            
                            cells = []
                            if len(col_highlight) != 0:
                                for c_idx in col_highlight:
                                    for r in range(rep_row+1, 30):
                                        cell_ = worksheet.cell(row=r, column=c_idx)
                                        if not cell_.value is None and not isinstance(cell_.value, str):
                                            cells.append([cell_.value, k, r])
                            if len(cells) != 0:
                                cells = sorted(cells, key = lambda x: x[0])
                                
                               
                                cells_[sn] = [cells[-1], cells[-2], cells[-3]]
                                if col == 'Қониқиш даражаси':
                                    cells_[sn] = [cells[1], cells[2], cells[3]]
                                
                                    

                    cell.alignment = alignment


            border = Border(
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000')
                        ) 

            cell_range = worksheet[f'A5:{get_column_letter(edge_col)}{len(ctab)+4}']
            for row in cell_range:
                for cell in row:
                    cell.border = border
                
            for nc in none_columns:
                worksheet.merge_cells(start_row=5, start_column=nc[1]+1, end_row=5, end_column=nc[0]+1)


            worksheet.print_options.horizontalCentered = True
            worksheet.print_options.verticalCentered = False
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
            worksheet.page_margins.top = 0.75
            worksheet.page_margins.bottom = 0.75
            worksheet.page_margins.left = 0.25
            worksheet.page_margins.right = 0.25
            worksheet.page_margins.header = 0.3
            worksheet.page_margins.footer = 0.3
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            workbook.save(out_filename)
            
    xl = openpyxl.load_workbook(out_filename)
    for i, sheet_name in enumerate(xl.sheetnames):
        i += 1
        worksheet = xl[sheet_name]
        
        worksheet.column_dimensions['A'].width = 36
        if i >=6:
            for row_idx in range(6, 21):
                worksheet.row_dimensions[row_idx].height = 40
            worksheet.row_dimensions[21].height = 50 
            worksheet.row_dimensions[6].height = 50 
            worksheet.row_dimensions[5].height = 125
            for c in range(1, 50):
                cell = worksheet.cell(row=6, column=c)
                cell.font = ft_column

            if i == 6:
                worksheet.column_dimensions['B'].width = 36
                worksheet.column_dimensions['C'].width = 36
                worksheet.column_dimensions['D'].width = 36
                worksheet.column_dimensions['E'].width = 36
                worksheet.column_dimensions['F'].width = 36
                worksheet.column_dimensions['G'].width = 36
            if i == 7:
                worksheet.column_dimensions['B'].width = 36
                worksheet.column_dimensions['C'].width = 36
                worksheet.column_dimensions['D'].width = 36
                worksheet.column_dimensions['E'].width = 36
                worksheet.column_dimensions['F'].width = 36
                worksheet.column_dimensions['G'].width = 36
            if i == 8:
                worksheet.row_dimensions[6].height = 60
                worksheet.column_dimensions['B'].width = 36
                worksheet.column_dimensions['C'].width = 36
                worksheet.column_dimensions['D'].width = 36
                worksheet.column_dimensions['E'].width = 36
                worksheet.column_dimensions['F'].width = 36
                worksheet.column_dimensions['G'].width = 36
                worksheet.column_dimensions['H'].width = 36


            
        else:
            worksheet.row_dimensions[5].height = 50
            for row_idx in range(6, 20):
                worksheet.row_dimensions[row_idx].height = 40
            if i == 1:
                worksheet.column_dimensions['D'].width = 34
                worksheet.column_dimensions['H'].width = 28
                worksheet.column_dimensions['I'].width = 22

        if i in [9, 10]:
            worksheet.row_dimensions[6].height = 100
            for c in range(2, 14):
                worksheet.column_dimensions[get_column_letter(c)].width = 25
            
            if i == 10:
                worksheet.column_dimensions['B'].width = 25



        if i in [11, 12, 13]:
            worksheet.row_dimensions[6].height = 100
            for c in range(2, 14):
                worksheet.column_dimensions[get_column_letter(c)].width = 22
                worksheet.row_dimensions[20].height = 50 
                if i in [11,12]:
                    worksheet.row_dimensions[6].height = 51
                elif i == 13:
                    worksheet.row_dimensions[6].height = 68
        
        if i in [14]:
            worksheet.row_dimensions[6].height = 113
            worksheet.column_dimensions['E'].width = 22
            worksheet.column_dimensions['K'].width = 22
            worksheet.column_dimensions['B'].width = 26
            worksheet.column_dimensions['B'].width = 23
            worksheet.column_dimensions['C'].width = 21
            worksheet.column_dimensions['D'].width = 21
            worksheet.column_dimensions['E'].width = 29
            worksheet.column_dimensions['F'].width = 21
            worksheet.column_dimensions['G'].width = 21
            worksheet.column_dimensions['H'].width = 21
            worksheet.column_dimensions['I'].width = 22
            worksheet.column_dimensions['J'].width = 21
            worksheet.column_dimensions['K'].width = 21
            worksheet.column_dimensions['L'].width = 40
            worksheet.row_dimensions[6].height = 73

        if i == 5:
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 40
            worksheet.column_dimensions['E'].width = 40
            worksheet.column_dimensions['F'].width = 40
            worksheet.column_dimensions['G'].width = 40
            worksheet.column_dimensions['H'].width = 40
            worksheet.row_dimensions[5].height = 145
        if i == 15:
            for c in range(2, 14):
                worksheet.column_dimensions[get_column_letter(c)].width = 26
            
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['I'].width = 33
            worksheet.row_dimensions[6].height = 75
            
        if i in [14, 15, 7, 6, 2]:
            for row_i in range(7, 22):
                worksheet.row_dimensions[row_i].height = 48
        
        if i in [10, 9]:
            for row_i in range(7, 22):
                worksheet.row_dimensions[row_i].height = 53
        
        # if i in [5]:
        #     for row_i in range(7, 22):
        #         worksheet.row_dimensions[row_i].height = 61.5
        
        
        if i in [15]:
            worksheet.row_dimensions[6].height = 85

        if i in list(range(1, 23)):
            # worksheet.column_dimensions['B'].width = 25
            # worksheet.column_dimensions['D'].width = 28
            # worksheet.column_dimensions['J'].width = 28
            worksheet.row_dimensions[20].height = 61.5
            worksheet.row_dimensions[21].height = 15
            worksheet.row_dimensions[22].height = 15
            worksheet.row_dimensions[23].height = 15
            worksheet.row_dimensions[24].height = 15
            worksheet.row_dimensions[25].height = 15
        
        if i in list(range(23, 100)):
            # worksheet.column_dimensions['B'].width = 25
            # worksheet.column_dimensions['D'].width = 28
            # worksheet.column_dimensions['J'].width = 28
            worksheet.row_dimensions[19].height = 61.5
            worksheet.row_dimensions[20].height = 15
            worksheet.row_dimensions[21].height = 15
            worksheet.row_dimensions[22].height = 15
            worksheet.row_dimensions[23].height = 15
            worksheet.row_dimensions[24].height = 15
            worksheet.row_dimensions[25].height = 15
        

        if i in [23, 24]:
            worksheet.row_dimensions[5].height = 257.25
            for letter_num in range(2,28):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21
        
        if  'Ер статистикаси' in sheet_name:
            worksheet.row_dimensions[5].height = 166.5
            for letter_num in range(2,5):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 33.29
        
        if  'Кредитга б' in sheet_name:
            worksheet.row_dimensions[5].height = 159.75
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 69.43
        
        if  'Харажатлар' in sheet_name:
            worksheet.row_dimensions[5].height = 122.25
            for letter_num in range(2,9):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 47.57
        
        if  '5.4. 2022 йилда ердан олинган д' in sheet_name:
            worksheet.row_dimensions[5].height = 176.25
            for letter_num in range(2,19):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 18.71
                if get_column_letter(letter_num) in ['F', 'G']:
                    worksheet.column_dimensions[get_column_letter(letter_num)].width = 24.86
        

        if  '4.7. Қандай қонунбу' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        
        if '4.4. Сизнингча, ерлардан унумли' in sheet_name:
            worksheet.row_dimensions[5].height = 168
            for letter_num in range(2,22):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 12.29
                if get_column_letter(letter_num) in ['J', 'K']:
                    worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.57
        

        if  '4.3. Фаолиятингизда қанд' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,30):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 12.29
        

        if  '4.2. Фаолиятингизга Қишлоқ хўжа' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,12):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
               
        if  '4.1. Ҳосилдорликни ошириш бўйич' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,12):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        

        if  '3.3. Қайси турдаги қишлоқ хўжал' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        
        if  '2.11. Сув таъминоти бил' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            worksheet.row_dimensions[6].height = 54.75
            for letter_num in range(2,12):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.86
        
        if  '2.10. Суғоришни' in sheet_name:
            worksheet.row_dimensions[5].height = 87.75
            worksheet.row_dimensions[6].height = 60.75
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 20.29
        

        if  '2.9. Ер майдонларини суғ' in sheet_name:
            worksheet.row_dimensions[5].height = 98.25
            worksheet.row_dimensions[6].height = 54
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.29
        

        if  '2.5. Ер олишда қандай' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,14):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.29
        

        if  '1.9. (АГАР ҲОКИМ ЁРДАМЧИСИ БУНИ' in sheet_name:
            worksheet.row_dimensions[5].height = 86.25
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.29
        

        if  '5.3. Ушбу ердан фойдаланиш уй х' in sheet_name:
            worksheet.row_dimensions[5].height = 61.5
            worksheet.row_dimensions[6].height = 49.5
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.86
        

        if  '5.1. Ушбу ер майдонларига эгали' in sheet_name:
            worksheet.row_dimensions[5].height = 68.25
            worksheet.row_dimensions[6].height = 49.5
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 20.71
        

        if  '4.5. Қўшимча молиявий ресурслар' in sheet_name:
            worksheet.row_dimensions[5].height = 71.25
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21
        
        if  '2.8. Ажрати' in sheet_name:
            worksheet.row_dimensions[5].height = 114
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 18.57
        

        if  '2.7. Ер ажратилган пайтд' in sheet_name:
            worksheet.row_dimensions[5].height = 90.75
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 17.14
        
        if  '2.4. Фойдаланаётган ер майдонин' in sheet_name:
            worksheet.row_dimensions[5].height = 72.75
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 20.86
        
        if  '2.3. Деҳқончилик қилиш, экинлар' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        

        if  '1.3. 2022 йилда ушбу ' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29

        
        if  '1.2. Респонде' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        


        if  '1.1.6. (БУНИ ЎҚИМАНГ) Респонден' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29

                    

        
            
        
        for i in range(5, 100):
            cell = worksheet.cell(row=i, column=1)
            cell.alignment = Alignment(horizontal='left', 
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=True,
                                        shrink_to_fit=True,
                                        indent=0) 
        

        


        
        

        
        
        
        xl.active = xl[sn]
        ws = xl.active
        
        

        
        
        

    xl.save(out_filename)
        # print(sheet_name)

    # wb = openpyxl.load_workbook(out_filename)
    # sheetNameList = wb.get_sheet_names()
    # for sheetName in sheetNameList:
    #     ws_active = wb.get_sheet_by_name(sheetName)
    #     cell = ws_active.cell(row=2, column=1)
    #     cell.value = get_rich_text(cell.value)
    #     if sheetName == 'Sheet5':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             cell.value = get_rich_text(cell.value, size=SIZE)
        
    #     if sheetName == 'Sheet6':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)

    #     if sheetName == 'Sheet7':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)
    #     if sheetName == 'Sheet9':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)
    #     if sheetName == 'Sheet10':
    #         for i in range(2, 15):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)


    #     ws_active.sheet_properties.pageSetUpPr.fitToPage = True
    #     ws_active.page_setup.fitToHeight = True
    #     ws_active.page_setup.page_margins = True

        
    #     if sheetName in list(cells_.keys()):
    #         ft = ft_blue
    #         for cell_element in cells_[sheetName]:
    #             if sheetName =='Sheet1':
    #                 cell_element[1] = 7
    #             if sheetName =='Sheet5':
    #                 cell_element[1] = 3
    #                 ft = ft_red
    #             cell = ws_active.cell(row=cell_element[2], column=cell_element[1])
    #             cell.font = ft
    
    # wb.save(out_filename)



    

    print(f'{datetime.today()} - Starting generating REGIONAL REPORT file...')

