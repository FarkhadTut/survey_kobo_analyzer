

from datetime import datetime
import datetime as Datetime
from download import excel_to_pandas
from default import URL
import pandas as pd 
from default import REPORT_COLS, outpath
import os 
import movecolumn as mc 
import numpy as np
import string
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import Border, Side
import re
from utils.utils import get_rich_text, SIZE, FONT_NAME, HEADER_SIZE, crosstab, getIndexes
import warnings
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

def convert_to_dict_of_lists(lst):
    result_dict = {}
    for item in lst:
        key = item[0]
        value = item[1]
        if key in result_dict:
            result_dict[key].append(value)
        else:
            result_dict[key] = [value]
    return result_dict

root = os.getcwd()
report_cols_list = [REPORT_COLS[key] for key in list(REPORT_COLS.keys())]
alphabet = dict.fromkeys(string.ascii_uppercase, 0)
alphabet = dict(zip(range(len(list(alphabet.keys()))), list(alphabet.keys())))


def to_excel(df, name):
    name = os.path.join(outpath, name)
    df.to_excel(name, index=False)







def generate_pdf():
    today = str(datetime.today().date()).replace('-', '_')
    db_filename = 'db_{today}.xlsx'.format(today=today)
    df = excel_to_pandas(URL, 'data/' + db_filename)
    return df


def analyze_suspic(df):
    today = str(datetime.today().date()).replace('-', '_')
    report_filename = 'suspicious_{today}.xlsx'.format(today=today)
    out_filename = f'out\\suspicious\\{report_filename}'
    COL = '8. Неча йилдан бери фаолият юритасиз?'
    mask = df[COL] > 32
    years = list(range(1991, 2024))
    for y in years:
        mask = (mask)&(df[COL] != y)

    COLUMNS = ['region', '3. Маҳаллангизни танланг?', 'label', COL]
    df_out = df[mask]
    # df_out = df_out.groupby(by=['region', '3. Маҳаллангизни танланг?', '2. Туманингизни танланг?'])['6. Респондентни танланг:'].count()
    df_out = df_out[COLUMNS]
    df_out.sort_values(by=['region', '3. Маҳаллангизни танланг?'], inplace=True, ascending=True)
    df_out.set_index(['region', '3. Маҳаллангизни танланг?'], inplace=True)
    df_out.rename(columns={'label': 'Респондент номи'}, inplace=True)

    with pd.ExcelWriter(out_filename, mode='w', engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name='Стаж (йил)')


    
    COL = '7. Ишчилар сони?'
    LIMIT_EMP = 1000
    mask = df[COL] > LIMIT_EMP

    COLUMNS = ['region', '3. Маҳаллангизни танланг?', 'label', COL]
    df_out = df[mask]
    # df_out = df_out.groupby(by=['region', '3. Маҳаллангизни танланг?', '2. Туманингизни танланг?'])['6. Респондентни танланг:'].count()
    df_out = df_out[COLUMNS]
    df_out.sort_values(by=['region', '3. Маҳаллангизни танланг?'], inplace=True, ascending=True)
    df_out.set_index(['region', '3. Маҳаллангизни танланг?'], inplace=True)
    df_out.rename(columns={'label': 'Респондент номи'}, inplace=True)

    with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'Ишчилар сони (> {LIMIT_EMP})')


    wb = openpyxl.load_workbook(out_filename)
    for sheet_name in wb.sheetnames:
        worksheet = wb[sheet_name]
        worksheet.column_dimensions['A'].width = 41.43
        worksheet.column_dimensions['B'].width = 41.43
        worksheet.column_dimensions['C'].width = 41.43
        if sheet_name == 'Стаж (йил)':
            worksheet.column_dimensions['D'].width = 38.14
        elif sheet_name == f'Ишчилар сони (> {LIMIT_EMP})':
            worksheet.column_dimensions['D'].width = 16.71

    wb.save(out_filename)

def report(df):
    mode ='w'
    prefix = ''
    today = str(datetime.today().date()).replace('-', '_')
    report_filename = 'report_analysis_{today}.xlsx'.format(today=today)
    out_filename = f'out/report/{report_filename}'
    total_col = 'Жами:'
    plan_col = "Режа"
    # df=pd.read_excel(f'out\\db_{today}.xlsx')
    df['_submission_time'] = pd.to_datetime(df['_submission_time']).dt.date

    # df.drop_duplicates(subset=['6. Респондентни танланг:', 'region', '2. Туманингизни танланг?'], keep='last', inplace=True)
    df_plan = pd.read_csv('mahalla_data.csv').reset_index()
    df_plan = df_plan[['region', 'district', 'mahalla', 'index']]
    df_plan.rename(columns={'index': plan_col}, inplace=True)
    df_orig = df.copy()

    
    df = df[['region', 'district', 'mahalla', '_id', '_submission_time']]
    df['_submission_time'] = df['_submission_time'].astype(str).str[:7]


    ########### REGIONS #####################
    df_ctab = pd.crosstab(index=[df['region']], 
                     columns=df['_submission_time'], 
                     margins=True,
                     margins_name=total_col)
    # df_mahalla_plan = df_plan.groupby(by=['region', 'district']).count()
    df_plan[plan_col] = plan_col
    df_mahalla_plan = pd.crosstab(index=[df_plan['region']], 
                     columns=df_plan[plan_col], 
                     margins=True,
                     margins_name=total_col)
    
    df_ctab = pd.merge(df_mahalla_plan, df_ctab, how='left', left_index=True, right_index=True, suffixes=('_plan', ''))
    df_ctab.drop(columns=[c for c in df_ctab.columns if '_plan' in c], inplace=True)
    

    df_ctab = mc.MoveToLast(df_ctab, plan_col)
    df_ctab['Ҳолати (%)'] = (df_ctab[total_col].div(df_ctab[plan_col])*100).round(1)
    df_ctab.fillna(0, inplace=True)
    with pd.ExcelWriter(out_filename, mode='w', engine='openpyxl') as writer:
        df_ctab.to_excel(writer, sheet_name='regions')
    ########################################

    ########### DISTRICTS #####################
    df_ctab = pd.crosstab(index=[df['region'], df['district']], 
                     columns=df['_submission_time'], 
                     margins=True,
                     margins_name=total_col)
    # df_mahalla_plan = df_plan.groupby(by=['region', 'district']).count()
    df_plan[plan_col] = plan_col
    df_mahalla_plan = pd.crosstab(index=[df_plan['region'], df_plan['district']], 
                     columns=df_plan[plan_col], 
                     margins=True,
                     margins_name=total_col)
    
    df_ctab = pd.merge(df_mahalla_plan, df_ctab, how='left', left_index=True, right_index=True, suffixes=('_plan', ''))
    df_ctab.drop(columns=[c for c in df_ctab.columns if '_plan' in c], inplace=True)
    

    df_ctab = mc.MoveToLast(df_ctab, plan_col)
    df_ctab['Ҳолати (%)'] = (df_ctab[total_col].div(df_ctab[plan_col])*100).round(1)
    df_ctab = pd.concat([df_ctab.tail(1), df_ctab.head(-1)], axis=0)

    df_ctab.fillna(0, inplace=True)
    with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
        df_ctab.to_excel(writer, sheet_name='districts')
    ########################################


    ########### MAHALLA #####################
    df_ctab = pd.crosstab(index=[df['region'], df['district'], df['mahalla']],
                     columns=df['_submission_time'], 
                     margins=True,
                     margins_name=total_col)
    # df_mahalla_plan = df_plan.groupby(by=['region', 'district']).count()
    df_plan[plan_col] = plan_col
    df_mahalla_plan = pd.crosstab(index=[df_plan['region'], df_plan['district'], df_plan['mahalla']], 
                     columns=df_plan[plan_col], 
                     margins=True,
                     margins_name=total_col)
    
    df_ctab = pd.merge(df_mahalla_plan, df_ctab, how='left', left_index=True, right_index=True, suffixes=('_plan', ''))
    df_ctab.drop(columns=[c for c in df_ctab.columns if '_plan' in c], inplace=True)
    

    df_ctab = mc.MoveToLast(df_ctab, plan_col)
    df_ctab['Ҳолати (%)'] = (df_ctab[total_col].div(df_ctab[plan_col])*100).round(1)
    df_ctab = pd.concat([df_ctab.tail(1), df_ctab.head(-1)], axis=0)
    df_ctab.fillna(0, inplace=True)

    with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl') as writer:
        df_ctab.to_excel(writer, sheet_name='mahalla')
    ########################################
   


    
    workbook = openpyxl.load_workbook(filename=out_filename)
    alignment = Alignment(horizontal='center', 
                vertical='center',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True,
                indent=0) 
    

    weird_green_clr = Color('ff92d050')
    green_clr = Color('ff00b04f')
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    for sn in workbook.sheetnames:
        worksheet = workbook[sn]
        
        if sn == 'districts':
            columns = ['A', 'B']
            for i, _ in enumerate(range(worksheet.max_column)):
                cell = worksheet.cell(row=1, column=i+1)
                filling = PatternFill(patternType='solid', fgColor=green_clr)
                cell.fill = filling

                cell = worksheet.cell(row=2, column=i+1)
                filling = PatternFill(patternType='solid', fgColor=weird_green_clr)
                cell.fill = filling

                # cell = worksheet.cell(row=worksheet.max_row, column=i+1)
                # filling = PatternFill(patternType='solid', fgColor=weird_green_clr)
                # cell.fill = filling

                for r, _ in enumerate(range(worksheet.max_row)): 
                    worksheet.cell(row=r+1, column=i+1).border = thin_border
                    if i > 0:
                        worksheet.cell(row=r+1, column=i+1).alignment = alignment
                    elif r == 0:
                        worksheet.cell(row=r+1, column=i+1).alignment = alignment
            worksheet.column_dimensions['B'].width = 21.86      
            
        elif sn == 'regions':
            columns = ['A']
            for i, _ in enumerate(range(worksheet.max_column)):
                cell = worksheet.cell(row=worksheet.max_row, column=i+1)
                filling = PatternFill(patternType='solid', fgColor=weird_green_clr)
                cell.fill = filling

                cell = worksheet.cell(row=1, column=i+1)
                filling = PatternFill(patternType='solid', fgColor=weird_green_clr)
                cell.fill = filling

                
                for r, _ in enumerate(range(worksheet.max_row)): 
                    worksheet.cell(row=r+1, column=i+1).border = thin_border
                    if i > 0:
                        worksheet.cell(row=r+1, column=i+1).alignment = alignment
                    elif r == 0:
                        worksheet.cell(row=r+1, column=i+1).alignment = alignment
                if i > 4 and i < worksheet.max_column-3:
                    worksheet.column_dimensions[get_column_letter(i+1)].width = 5
                elif i >= worksheet.max_column-3:
                    worksheet.column_dimensions[get_column_letter(i+1)].width = 7.43

                    

        elif sn == 'mahalla':
            columns = ['A', 'B', 'C']

        for i, c in enumerate(columns):
            worksheet.column_dimensions[c].width = 30
            
    

         
    workbook.save(filename=out_filename)
    print(df_ctab)

    exit()
    
    return out_filename







def report_all(df, *args, **kwargs):
    mode ='w'
    prefix = ''
    today = str(datetime.today().date()).replace('-', '_')
    report_filename = 'all_report_analysis_{today}.xlsx'.format(today=today)
    out_filename = f'out/report/{report_filename}'
    # df=pd.read_excel(f'out\\db_{today}.xlsx')
    # df.drop_duplicates(subset=['name'], keep='last', inplace=True)
    df.drop_duplicates(subset=['region', '2. Туманингизни танланг?', '3. Маҳаллангизни танланг?'], keep='last', inplace=True)

    df['_submission_time'] = pd.to_datetime(df['_submission_time']).dt.date
    df_orig = df.copy()
    df.rename(columns={'region': 'region', '2. Туманингизни танланг?': 'district',}, inplace=True)
    df = df[['region', 'district', '_id', '_submission_time']]

    
    dates = df_orig['_submission_time'].unique()
    df_out = pd.DataFrame()
    
    print(pd.crosstab(df, values='_id', columns=['region', ]))

    workbook = openpyxl.load_workbook(filename=out_filename)
    alignment = Alignment(horizontal='left', 
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True,
                indent=0) 
    for sn in workbook.sheetnames:
        worksheet = workbook[sn]
        if sn == 'districts':
            columns = ['A', 'B']
            
        elif sn == 'regions':
            columns = ['A']
        elif sn == 'mahalla':
            columns = ['A', 'B', 'C']

        for i, c in enumerate(columns):
            worksheet.column_dimensions[c].width = 30
         
    workbook.save(filename=out_filename)




    workbook = openpyxl.load_workbook(filename=filename_nd)
    alignment = Alignment(horizontal='left', 
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=True,
                indent=0) 
    for sn in workbook.sheetnames:
        worksheet = workbook[sn]
        columns = ['A', 'B', 'C']

        for i, c in enumerate(columns):
            worksheet.column_dimensions[c].width = 30
         
    workbook.save(filename=filename_nd)
    return report_filename


def highlight(x):
    print(x)
    return x

def freq_table(df):
    
    # if 'out' not in df_filename:
    #     df_filename = os.path.join(root, 'out', df_filename)
    # df = pd.read_excel(df_filename)
    df.dropna(axis=1, inplace=True, how='all')
    df_orig = df.copy()

    # df.drop_duplicates(subset=['6. Респондентни танланг:', 'region', '2. Туманингизни танланг?'], keep='first', inplace=True)
    
    
    today = str(datetime.today().date()).replace('-', '_')

    start_col = 'region'
    end_col = '44. Бошқа таклиф ва мулоҳазалар (Ҳоким ёрдамчисига қўшимча ваколатлар бериш, қўшимча вазифа ва функциялар юклаш (ёки олиб ташлаш)):'

    columns = df.columns.values.tolist()
    columns = columns[columns.index(start_col) : columns.index(end_col)+1]

    columns_temp = []
    for i,c in enumerate(columns):
        if i + 1 < len(columns):
            if not c + '/' in columns[i+1] and not c + ' /' in columns[i+1]:
                columns_temp.append(c)
        else:
            columns_temp.append(c)
    
    text_cols =  ["44. ", ]

    for i, c in enumerate(columns_temp[:]):
        for text_col in text_cols:
            if c.startswith(text_col) or '(Бошқа)' in c:
                if c in columns_temp:
                    del columns_temp[columns_temp.index(c)]
                

    columns = columns_temp
    all_columns = columns
    # columns = ['region',
    #            '2. Туманингизни танланг?',
    #            ]

    df = df[columns]
    df.reset_index(drop=True, inplace=True)
    
    columns = df.columns
    df_out = pd.DataFrame()

    filename_out = f'out/freq/hy_freq_table_{today}.xlsx'
    df_big = pd.DataFrame()

    # town_vil = [ 'Шаҳар', 'Қишлоқ', 'Аёл', 'Эркак', '18-35', '36-49', '50-999','all',]
    town_vil = ['all',]
    for i, c in enumerate(list(reversed(df.columns.values))):

        
        df_temp = pd.DataFrame()
        for j, tv in enumerate(town_vil):
            if tv in ['Шаҳар', 'Қишлоқ']:
                mask = df_orig['1. САВОЛ БЕРМАЙ, Қаерда (Шаҳар ёки қишлоқда) яшашини киритинг?'] == tv
                df = df_orig[mask]
            elif tv in ['18-35', '36-49', '50-999']:
                btm_age, top_age = tv.split('-')
                btm_age, top_age = int(btm_age), int(top_age)
                if btm_age == 18:
                    btm_age = 17
                btm_mask = df_orig['6. Ёшингиз?'] >= btm_age
                top_mask = df_orig['6. Ёшингиз?'] <= top_age
                # mask = ~(btm_mask & top_mask)
                df = df_orig[btm_mask][top_mask]
                df = df.dropna(subset='6. Ёшингиз?', axis=0)
                
            elif tv in ['Аёл', 'Эркак']:
                mask = df_orig['2. САВОЛ БЕРМАЙ, респондентларнинг жинсини аниқланг:'] == tv
                df = df_orig[mask]
            else:
                df = df_orig.copy()
            
            ctab = pd.crosstab(index = df.index, columns = df[c])
            ctab = ctab.sum()
            

            ctab_pct = (ctab/ctab.values.sum()).round(decimals=4)
            
            tab = pd.concat([ctab, ctab_pct], axis=1)
            
            cols = ['Frequency', 'Percentage (%)', c]
            tab[c] = tab.index
            
            
            df_out = pd.DataFrame(columns=cols, data=tab.values)
            df_out = mc.MoveTo1(df_out, c)
            

            if not df[c].dtype == 'object':
                df_out['Average score'] = df_out['Percentage (%)'] * df_out[c]
                df_out['Average score'] = df_out['Average score'].round(decimals=3)
            else:
                df_out['Average score'] = [np.nan] * len(df_out['Percentage (%)'])
            
            
            fillna_cols = df_out.columns.values.tolist()
            
            df_out[['Percentage (%)']] = df_out[['Percentage (%)']] * 100
            

            
            freq_sum = df_out['Frequency'].sum()
            pct_sum =df_out['Percentage (%)'].sum()
            avg_sum = df_out['Average score'].sum()
            if avg_sum == 0:
                avg_sum = np.nan
            df_out['Frequency'] = df_out['Frequency'].astype(float).astype('Int64')
            df_total = pd.DataFrame(columns=df_out.columns, data=[['Total:', freq_sum,\
                                                                pct_sum,
                                                                 avg_sum]])
            
            
            df_out.sort_values(by=['Percentage (%)'], ascending=False, inplace=True)
            df_out = pd.concat([df_out, df_total], axis=0)
       
            

            
            # Add space after tables
            df_out.rename(columns={'Frequency': f'{tv}_Frequency', 'Percentage (%)': f'{tv}_Percentage (%)', 'Average score': f'{tv}_Average score'}, inplace=True)
            if j == 0:
                df_temp = df_out
            else:
                if tv == 'all' or tv in ['Аёл'] or tv == '18-35':
                    space = pd.DataFrame(np.nan, index=list(range(len(df_out))), columns=['A', 'B'])
                    # space = df_out[df_out.columns.values[:2]].fill(0)
                    df_out.reset_index(drop=True, inplace=True)
                    df_out = pd.concat([space, df_out], axis=1)
                
         
                df_temp = pd.merge(df_temp, df_out, how='outer', on=c)
                fillna_cols = df_temp.columns.values.tolist()
                fillna_cols = [fc for fc in fillna_cols if fc != 'Average score' and fc != c and fc not in ['A', 'B'] and 'A_' not in fc and 'B_' not in fc]
                df_temp[fillna_cols] = df_temp[fillna_cols].fillna(0)
            

        df_space = pd.DataFrame(columns=df_temp.columns.values.tolist(), data=[[np.nan]*len(df_temp.columns),
                                                                                [np.nan]*len(df_temp.columns)])
        
        ### move the Total row to the last
        col_first = df_temp.columns.values[0]
        mask = df_temp[col_first] == 'Total:'
        df_temp = pd.concat([df_temp[~mask], df_temp[mask]], axis=0)

        ## fill nan values with zeroes
        # df_temp.fillna(0, inplace=True)

        ## add space between question blocks
        df_temp = pd.concat([df_temp, df_space], axis=0)

        

        # columns = [c] + ['Frequency_x', 'Frequency_y', 'Percentage (%)_x',  'Percentage (%)_y', 'Average score_x', 'Average score_y', 'A', 'B', 'Frequency', 'Percentage (%)', 'Average score']
        # print(df_temp.columns.values)
        # sys.exit()
        columns = [c] + [
                          'all_Frequency', 'all_Percentage (%)', 'all_Average score']
        


        

        micolumns = pd.MultiIndex.from_tuples(
                                            [(None, c),
                                             ('Across all data', "Frequency"), 
                                             (None, "Percentage (%)"),
                                             (None, "Average score")],)
        
        # columns = [c] + ['all_Frequency', 'all_Percentage (%)', 'all_Average score']
   
        # micolumns = pd.MultiIndex.from_tuples(
        #                                     [(None, c),
        #                                      ('Across all data', "Frequency"), 
        #                                      (None, "Percentage (%)"),
        #                                      (None, "Average score")
        #                                      ],)

        df_temp = df_temp[columns]
        df_temp = pd.DataFrame(columns=micolumns, data=df_temp.values)

    
        
        
 
        if i == 0:
            df_big = df_temp
        else:
            
            df_temp = df_temp.reset_index(drop=True).T.reset_index().T
            df_temp.rename(columns=dict(zip(df_temp.columns.values.tolist(), df_big.columns.values.tolist())), inplace=True)
            if i == 1:
                columns = df_big.columns.values.tolist()
                df_big = df_big.reset_index(drop=True).T.reset_index().T
                df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), columns)), inplace=True)
            
            df_big = pd.concat([df_temp, df_big], axis=0)

    
        

    new_columns = [np.nan] * len(df_big.columns.values)
    new_columns[0] = 'Hokim yordamchilari 2024 Frequency table'

    df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), new_columns)), inplace=True)
    
    df_big.reset_index(drop=True, inplace=True)
    with pd.ExcelWriter(filename_out, mode='w', engine='openpyxl') as writer:
        df_big.to_excel(writer)


    wb = openpyxl.load_workbook(filename_out)
    worksheet = wb.active
    worksheet.delete_cols(1)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    worksheet.column_dimensions['A'].width = 35.57
    alignment = Alignment(horizontal='left', 
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=True,
                                    shrink_to_fit=True,
                                    indent=0)
    for c in all_columns:
        rowcol = getIndexes(df_big, c)
        row = rowcol[0][0]
        cell = worksheet.cell(row=row+2, column=1)
        cell.alignment = alignment



    wb.save(filename_out)
    print(f'{datetime.today()} - Starting generating FREQ TABLE file...')
    return filename_out
    

def report_by_region(df):
    # df.rename(columns={'33.  Сизнингча, (вилоят ҳокими) вилоят/республика ҳокими сифатида қандай ишлаяпти?': '33. Вилоят хокими қандай ишлаяпти?',
    #                    '36. Сизнингча, туманингиз ҳокими ўз лавозимида қандай ишламоқда?': '36. Туман ҳокими ўз лавозимида қандай ишламоқда?',
    #                    "40. Ўзбекистон Республикаси Олий Мажлис Қонунчилик палатасининг амалдаги таркибининг фаолиятини қандай баҳолайсиз?": "40. Олий Мажлис Қонунчилик палатасининг амалдаги таркибининг фаолиятини қандай баҳолайсиз?",
    #                    "7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?": "7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангиз?",
    #                    "8. Вилоятингиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?": "8. Вилоятингиздаги умумий вазиятдан қониқиш даражангиз?",
    #                    "10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?": "10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангиз?",
    #                    "7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?": "7.2. Мамлакатдаги умумий вазият",
    #                    "9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?":"9. Вилоятдаги умумий вазият",
    #                    "11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?":"11. Маҳалладаги умумий вазият",}, inplace=True)

    today = str(datetime.today().date()).replace('-', '_')
    out_filename = f'out/regional/regional_{today}.xlsx'
    
    df.replace({"Қониқарсиз\n(Сифатсиз хизматлар, бюрократия, коррупция ва б.)": "Қониқарсиз"}, regex=False, inplace=True)
    # cols = ['7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?',
    #         '7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?',
    #         '11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?',
    #         '9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?',
    #         '12. Сизнингча, атрофингиздаги одамлар ҳозир қандай кайфиятда: кўтаринки, хотиржам ёки тушкун (хавотирли)?',
    #         '13. “Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:',
    #         '16. “Ҳукумат фуқаролар билан очиқ мулоқотда бўлмоқда ва уларнинг муаммоларига ўз вақтида жавоб қайтармоқда" Мазкур фикрга...:',
    #         '14. “Мамлакатимизда олиб борилаётган ислоҳотлар тўғри йўлда кетмоқда”, мазкур фикрга:',
    #         '26. Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?',
    #         '27. Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?',
    #         '30. Сизнингча, Шавкат Мирзиёев мамлакатдаги вазиятни яхши томонга ўзгартира оладими ёки йўқми?']
    
    # q_nine = [c for c in df.columns.values.tolist() if not '/' in c and c.split(' ', maxsplit=1)[0].rstrip('.') in ['9.1','9.2','9.3','9.4','9.5','9.6','9.7','9.8','9.9','9.10','9.11','9.12','9.13','9.14','9.15','9.16','9.17','9.18','9.19','9.20','9.21','9.22','9.23']]

    cols = {
        'single': ['11. Экспорт қилаётган маҳсулотингизни ўзингиз ишлаб чиқарасизми?'],
        'multiple': ['9. Қайси соҳада экспорт фаолияти билан шуғулланасиз?'],
        'multicolumn': ['Фаолиятингиз давомида ҳамкорлик қиладиган ташкилотларнинг фаолиятини қандай баҳолайсиз?', 
                        'Корхона фаолияти тўғрисида',
                      
                        ],
        

    }
    region_col = 'region'
    df_out = pd.DataFrame()
    merge_columns = {}
    dfs = []
    sheet_names = []
    counter = 0
    cells_ = {}


    for type_idx, (type_, columns_) in enumerate(cols.items()):
        for sheet_idx, col in enumerate(columns_):
            if type_ == 'single':
                ctab = crosstab(index=df[region_col], columns=df[col])

            elif type_ == 'multiple':
                if col == ' балл':
                    columns = [
                        '25. Президент Шавкат Мирзиёев ўз лавозимида фикрингиз бўйича қандай фаолият кўрсатаётганлигини 7 баллик шкалада баҳоланг?',
                        '33. Вилоят хокими қандай ишлаяпти?',
                        "36. Туман ҳокими ўз лавозимида қандай ишламоқда?",
                        "40. Олий Мажлис Қонунчилик палатасининг амалдаги таркибининг фаолиятини қандай баҳолайсиз?",
                        "7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангиз?",
                        "8. Вилоятингиздаги умумий вазиятдан қониқиш даражангиз?",
                        "10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангиз?"
                    ]
                                    
                    
                else:
                    columns = [c for c in df.columns.values if col in c][1:]
                    
                    drop_by_key = ['Ҳеч бири',
                                   "Бошқа",
                                   "Жавоб бериш",
                                   "Ҳеч қандай",
                                   "Телефон алоқаси билан боғлиқ муаммо",
                                   'Конвертация очилганлиги (валюта сиёсати)']
                    for key in drop_by_key:
                        columns = [c for c in columns if not key in c]
                    
                    if col == '28. Илтимос, Шавкат Мирзиёевнинг Президентлик даврида эришган асосий ютуқларини номлаб беринг?':
                        prepend_list = ['Янги Конституция қабул қилиниши',
                                 'Сўз эркинлигининг кучайиши',
                                 'Ёшларни қўллаб-қувватлашнинг кучайтирилиши',
                                 'Дин ва эътиқод эркинлигининг кучайиши',
                                 'Тадбиркорликни қўллаб-қувватлаш ишлари',
                                 'Камбағалликка қарши кураш ишлари',
                                 'Халқ билан мулоқотнинг кучайиши',
                                 'Соғлиқни сақлаш тизимидаги ислоҳотлар',
                                 ]
                                 
                        columns_prepend = []
                        
                        for p in prepend_list:
                            for i, c in enumerate(columns[:]):
                                if p in c:
                                    columns_prepend.append(c)
                                    columns.pop(i)
                                    break
                        columns = columns_prepend + columns
                        
                dfs_multiple = {}
                for col_ in columns: 
                    # ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col_], normalize='index', margins=True) * 100).round(decimals=1)
                    # index = ctab_pct.index.values.tolist()
                    # index[-1] = 'Республика бўйича'
                    # ctab_pct.index = index
                    # ctab_pct = pd.concat([ctab_pct.tail(1), ctab_pct.head(-1)], axis=0)
                    # ctab = ctab_pct
                    
                    # # total_by_answers_pct['Жами'] = None

                    # col_name = col_.replace(col+'/', '') ## - WORKING WELL DO NOT DELETE THIS
                    if col+'/' in col_:
                        col_name = col_.split(col+'/')[-1] ## - EVEN BETTER
                    else:
                        col_name = col_.split(col+' /')[-1] ## - EVEN BETTER

                    print(col_name)
                    
                    ctab = crosstab(index = df[region_col], columns = df[col_], rename={1:col_name}, remove_names=False)
                    if col == ' балл':
                        col_name = col_name.split(' ', maxsplit=1)[1]

                        column_values = pd.Series(ctab.columns.values.tolist(), index=ctab.columns)
                        ctab = ctab/100
                        ctab = ctab.multiply(column_values, axis=1)
                        ctab['Average score'] = ctab.sum(axis=1)
                        ctab = ctab['Average score'].to_frame().round(1)
                        ctab.rename(columns={'Average score': 1}, inplace=True)

                    if not (col_name,    '%') in ctab.columns:
                        ctab.columns = ctab.columns.set_levels([col_name, None], level=0)
                        ctab[(col_name,    '%')] = 0
                    ctab.rename(columns={1:col_name}, inplace=True)
                    
                    # dfs_multiple.append(ctab[col_name])#.to_frame())
            
                    dfs_multiple[col_name] = ctab[col_name]#.to_frame())
                
                
                ctab = pd.concat(dfs_multiple.values(), axis=1, keys=dfs_multiple.keys())
            elif type_ == 'multicolumn':
                # columns = [c for c in df.columns.values if col in c]
                if col == 'Фаолиятингиз давомида ҳамкорлик қиладиган ташкилотларнинг фаолиятини қандай баҳолайсиз?':
                    columns = [c for c in df.columns.values if ('row' in c) and ('/' in c)]

                elif col == 'Корхона фаолияти тўғрисида':
                    columns = ['7. Ишчилар сони',
                               '8. Неча йилдан бери фаолият юритасиз?',
                               ]
                
                elif col == 'Солиқ тизимига оид саволлар':
                    columns = ['10. Ўзбекистонда тадбиркорлар учун солиқ юкини (ставкалари) қандай баҳолайсиз? ',
                               '12. ҚҚС ни қайтариш бўйича муаммога дуч келганмисиз?',
                               '12.1. ҚҚС ни қайтариш қанча вақт талаб этган?',
                               '13. 2022-2023 йилларда корхонангиз фаолиятида солиқ текширувлари (режали ва режадан ташқари) ўтказилганми?',
                               '14. Солиқ текшируви натижалари бўйича жавобгарликка тортиш чоралари (молиявий, маъмурий, жиноий санкциялар) кўрилганми? ',
                               '15. Сизнинг фикрингизча, ушбу солиқ текшируви натижалари бўйича жавобгарликка тортиш чоралари асосли ва адолатли бўлганми?',
                               '16. Ҳамкорларингиз (таъминотчи ва буюртмачи)нинг содир этган ҳуқуқбузарлиги (шубҳали шартнома ва транзакциялар) сабабли, Сиз ҳам солиқ органлари томонидан жаримага тортилганмисиз (ёки қизил йўлакка тушиб қолганмисиз)?',
                               '18. Сизнингча ҳамкорлар (таъминотчи ва буюртмачи) томонидан содир этилган солиқ ҳуқуқбузарлиги (шубҳали шартнома ва транзакциялар) учун ким жаримага тортилиши лозим?',
                               '19. Ер солиғи ставкасини ҳисоблашда қандай муаммолар бор?',
                               '20. Солиқ қонунчилигидан хабардорлик даражангизни қандай баҳолайсиз? ',
                               '21. Солиқ қонунчилигининг мураккаблик даражасини баҳоланг?',
                               '1-22. Солиқ тўлаш жараёни бўйича тажрибангизга таяниб, қуйидаги фикрларга муносабатингизни билдиринг?/Солиқ ходимлари профессионал ва холис',
                               '2-22. Солиқ тўлаш жараёни бўйича тажрибангизга таяниб, қуйидаги фикрларга муносабатингизни билдиринг?/Солиқ ходимлари солиқ тизимига оид маълумотларни батафсил ва шаффоф тақдим этади',
                               '3-22. Солиқ тўлаш жараёни бўйича тажрибангизга таяниб, қуйидаги фикрларга муносабатингизни билдиринг?/Солиқ ходимлари шикоят ва таклифларни эшитади',
                               ],
                               
          
                
                
                if '“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:' in columns:
                    col_idx = columns.index('“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:')
                    columns[col_idx] = '“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:'
                
                new_columns = [(c.split(' ', maxsplit=1)[1].replace('...', '')).strip() if not 'row' in c \
                               else (c.split('/', maxsplit=1)[1].replace('...', '')).strip() \
                               for c in columns]
                for c_idx, c in enumerate(new_columns):
                    if c == 'Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?':
                        new_columns[c_idx] = 'Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими?'
                    elif c == 'Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?':
                        new_columns[c_idx] = 'Шавкат Мирзиёевга ишонасизми?'
                    
                    elif 'Фаолиятингиз давомида ҳамкорлик қиладиган ташкилотларнинг фаолиятини қандай баҳолайсиз?' in c and '/' in c:
                        new_c = c.split('/', maxsplit=1)[1].strip()
                        new_columns[c_idx] = new_c
                        
                    
                    
                
                df.rename(columns=dict(zip(columns, new_columns)), inplace=True)
                columns = new_columns
                dfs_multiple = []
                data = {}

                for col_ in columns:
                    if not col in ['Корхона фаолияти тўғрисида']:
                        ctab = pd.crosstab(index = df[region_col], columns = df[col_])
                        ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col_], normalize='index', margins=True) * 100).round(decimals=1)
                        index = ctab_pct.index.values.tolist()
                        index[-1] = 'Республика бўйича'
                        ctab_pct.index = index
                        ctab_pct = pd.concat([ctab_pct.tail(1), ctab_pct.head(-1)], axis=0)
                        ctab = ctab_pct
                        col_name = col_.replace(col+'/', '')
                        print(col_)
                        if col == 'Фаолиятингиз давомида ҳамкорлик қиладиган ташкилотларнинг фаолиятини қандай баҳолайсиз?':
                            # if col_ == '15. Экспорт фаолиятингизни амалга ошириш давомида ҳамкорлик қиладиган ташкилотларнинг фаолиятини баҳоланг':
                            #     target_cols = ['Ишим тушмаган', 'Қониқарли', 'Қониқарсиз']
                            # else:
                            
                            target_cols = ['Ишим тушмаган', 'Қониқарли', 'Қониқарсиз']
                        data = {}
                        for ti, target_col in enumerate(target_cols):
                            if not target_col in ctab.columns:
                                data[(col_, target_col)] = [0] * len(ctab)
                            else:
                                data[(col_, target_col)] = ctab[target_col].values.tolist()

                        ctab = pd.DataFrame(data, index=ctab.index)
                    else:
                        data = df.groupby(by=['region'], as_index=True)[['region', col_]].mean()
                        total_data = (df[col_].mean())
                        total_data= pd.DataFrame(columns=['region', col_], data=[['Республика бўйича', total_data]])
                        total_data.set_index('region', inplace=True)
                        data = pd.concat([total_data, data], axis=0)

                        # if data.mean().values[0] < 200:
                        #     data = data.round(1)
                        # else:
                        #     data = data.round(0)
                        
                        ctab = pd.DataFrame(data, index=data.index)
                
                    dfs_multiple.append(ctab)
                    
                
                  
                ctab = pd.concat(dfs_multiple, axis=1)
                
        
            # ctab = (pd.crosstab(index = df[region_col], columns = df[col], normalize='index') * 100).round(decimals=1)
            ctab[col] = ctab.index 
            ctab = mc.MoveTo1(ctab, col)
            ctab = ctab.reset_index(drop=True).T.reset_index().T      
            ctab.reset_index(drop=True, inplace=True)
            if type_ != 'multicolumn':
                for ci in ctab.columns:
                    if ci != 0 and ci%2 == 0:
                        if ctab.at[0, ci] == ctab.at[0, ci-1]:
                            ctab.at[0, ci] = None
            
            # ctab.style.background_gradient(cmap='coolwarm').set_precision(2)
            # ctab.show()
            merge_columns[ctab.iloc[0][0]] = len(ctab.columns)

            dfs.append(ctab)
            
            sn = col
            if len(col) > 31:
                sn = col[:-31]
            if counter == 0:
                sn = sn.replace(':', '')
                with pd.ExcelWriter(out_filename, mode='w', engine='openpyxl') as writer:
                    ctab.T.reset_index().T.to_excel(writer, index=False, header=None, sheet_name=sn)
            else:
                with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl', if_sheet_exists='error') as writer:
                    ctab.T.reset_index().T.to_excel(writer, index=False,header=None, sheet_name=sn)
            
         
            counter +=1
            sheet_names.append(sn)
            
            # removing question nums from headers
            if col[0].isnumeric():
                col = col.split(' ', maxsplit=1)[1] 
            workbook = openpyxl.load_workbook(filename=out_filename)
            worksheet = workbook[sn]
            worksheet.sheet_view.zoomScale = 55
            ft_default = Font(size=SIZE, name=FONT_NAME)
            ft_blue = Font(color="0070c0", bold=True, size=SIZE, name=FONT_NAME)
            ft_red = Font(color="C00000", bold=True, size=SIZE, name=FONT_NAME)
            ft_column = Font(color='1f4e78', bold=True, size=SIZE, name=FONT_NAME)
            ft_header = Font(color='1f4e78', bold=True, size=HEADER_SIZE, name=FONT_NAME)
            ft_tiny = Font(color='1f4e78', italic=True, size=14, name=FONT_NAME)
            ft_bold = Font(color='000000', bold=True, italic=False, size=SIZE, name=FONT_NAME)
            edge_col = len(ctab.columns)
            for c in range(edge_col):
                cell = worksheet.cell(row=1, column=c+1)
                cell.value = None
            cell = worksheet.cell(row=1, column=1)
     


            cell.value = col
          
            
            cell.font = ft_header
            cell.alignment = Alignment(horizontal='center', 
                                vertical='center',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0) 
            worksheet.row_dimensions[2].height = 50
            worksheet.insert_rows(1)
            worksheet.insert_rows(3)
            worksheet.insert_rows(3)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=edge_col)

            
            none_columns = [(column_i, column_i-1) for column_i, column_ in enumerate(ctab.iloc[0].values.tolist()) if column_ is None]
            
            
            
            cell = worksheet.cell(row=5, column=1)
            cell.value = None
            # worksheet.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)

            worksheet.column_dimensions['A'].width = 34
            # cell = worksheet.cell(row=1, column=edge_col)
            # cell.value = f'{counter}-жадвал'

            cell = worksheet.cell(row=4, column=edge_col)
            if type_ == 'multicolumn':
                cell.value = f'(ўртача қиймат)'
            else:
                cell.value = f'(фоизда)'

            cell.alignment = Alignment(horizontal='right', 
                                vertical='bottom',
                                text_rotation=0,
                                wrap_text=False,
                                shrink_to_fit=False,
                                indent=0) 

            
            cell.font = ft_tiny

            if type_ == 'multicolumn':
                worksheet.row_dimensions[6].height = 60
                columns = ctab.iloc[0].tolist()[1:]
                all_cols = ctab.iloc[1].values.tolist()[1:]
                columns = list(zip(columns, all_cols))
                columns = convert_to_dict_of_lists(columns)
                key_count = 1
                for _, col_ in enumerate(list(columns.keys())):
                    target_cols = columns[col_]
                    for ti, target_col in enumerate(target_cols):
                        key_count += 1
                        if ti == len(target_cols)-1:
                            worksheet.merge_cells(start_row=5, start_column=key_count-len(target_cols)+1, end_row=5, end_column=key_count)
                            worksheet[f'{get_column_letter(key_count+1-len(target_cols))}5'] = col_
                            break
                        
                        cell = worksheet.cell(row=5, column=key_count)
                        cell.value = None
                        
                    
                    # worksheet[f'{alphabet[key_count-1]}5'] = col_

                    cell = worksheet.cell(row=6, column=key_count)
                    cell.font = ft_column

                worksheet.row_dimensions[6].height = 60
            
            for j in range(5, 22):
                alignment = Alignment(horizontal='general', 
                                vertical='bottom',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0) 
                cell = worksheet.cell(row=j, column=1)
                cell.alignment = alignment
                worksheet.row_dimensions[j].height = 30
                if j == 5:
                    worksheet.row_dimensions[j].height = 120
                    
                for k in range(1, edge_col+1):
                    cell = worksheet.cell(row=j, column=k)
                    cell.font = ft_default
                    if k == 1 and j == 6:
                        cell.font = ft_column
                        continue
                    if k != 1:
                        alignment.horizontal = 'center' 
                        alignment.vertical = 'center' 
                    if j == 5:
                        if len(ctab.columns) > 11:
                            if not col in ['Фаолиятингиз давомида ҳамкорлик қиладиган ташкилотларнинг фаолиятини қандай баҳолайсиз?',
                                           '5. ТАРМОҚНИ ТАНЛАНГ']:
                                alignment.text_rotation = 90
                                worksheet.row_dimensions[j].height = 190
                            worksheet.column_dimensions[get_column_letter(k)].width = 13
                        else:
                            worksheet.column_dimensions[get_column_letter(k)].width = 20
                            

                        cell.font = ft_column
                    
                    
                    else:  
                        rep_cell = worksheet.cell(row=6, column=1)
                        rep_vals = []
                        rep_row = None
                        resp_mask = ctab[0] == 'Республика бўйича'
                        resp_idx = ctab[resp_mask].index.values[0]+1+4
                       
                        resp_idx_ctab = ctab[resp_mask].index.values[0]
                        resp_values = []
                        for __c__ in ctab.columns:
                            if ctab.at[resp_idx_ctab-1, __c__] != 'Сони':
                                if not isinstance(ctab[resp_mask][__c__].values[0], str):
                                    resp_values.append(ctab[resp_mask][__c__].values[0])
                        

                        resp_mean = np.mean(resp_values)
                        thresh = resp_mean
                        lower_thresh = resp_mean
                        ft_high = ft_blue
                        ft_low = ft_red
                        
                        
                            
                        
                        if rep_cell.value == 'Республика бўйича':
                            rep_cell.font = ft_column
                            rep_vals = [ (worksheet.cell(row=6, column=c_).value, c_)  for c_ in range(1, edge_col+1) if worksheet.cell(row=5, column=c_).value != 'Сони'][1:]
                            rep_vals = [v[1] for v in rep_vals if v[0] >=thresh]
                            rep_row = 6
                        elif worksheet.cell(row=7, column=1).value == 'Республика бўйича':
                            rep_cell = worksheet.cell(row=7, column=1)
                            rep_cell.font = ft_column
                            rep_vals = [(worksheet.cell(row=7, column=c_).value, c_)  for c_ in range(1, edge_col+1) if worksheet.cell(row=6, column=c_).value != 'Сони'][1:]
                            
                            rep_vals = [v[1] for v in rep_vals if v[0] >=thresh]
                            rep_row = 7
                            
                        
                        
                        if not cell.value is None:
                            col_highlight = []
                            if not isinstance(cell.value, str):
                                if j  == rep_row:
                                    cell.font = ft_bold
                                if k in rep_vals:
                                    if cell.value >= thresh:
                                        cell.font = ft_high
                                    elif cell.value <=lower_thresh:
                                        cell.font = ft_low
                                                                            
                    
                            
                            cells = []
                            if len(col_highlight) != 0:
                                for c_idx in col_highlight:
                                    for r in range(rep_row+1, 30):
                                        cell_ = worksheet.cell(row=r, column=c_idx)
                                        if not cell_.value is None and not isinstance(cell_.value, str):
                                            cells.append([cell_.value, k, r])
                            if len(cells) != 0:
                                cells = sorted(cells, key = lambda x: x[0])
                                
                               
                                cells_[sn] = [cells[-1], cells[-2], cells[-3]]
                                if col == 'Қониқиш даражаси':
                                    cells_[sn] = [cells[1], cells[2], cells[3]]
                                
                                    

                    cell.alignment = alignment


            border = Border(
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000')
                        ) 

            cell_range = worksheet[f'A5:{get_column_letter(edge_col)}{len(ctab)+4}']
            for row in cell_range:
                for cell in row:
                    cell.border = border
                
            for nc in none_columns:
                worksheet.merge_cells(start_row=5, start_column=nc[1]+1, end_row=5, end_column=nc[0]+1)


            worksheet.print_options.horizontalCentered = True
            worksheet.print_options.verticalCentered = False
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
            worksheet.page_margins.top = 0.75
            worksheet.page_margins.bottom = 0.75
            worksheet.page_margins.left = 0.25
            worksheet.page_margins.right = 0.25
            worksheet.page_margins.header = 0.3
            worksheet.page_margins.footer = 0.3
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            workbook.save(out_filename)
            
    xl = openpyxl.load_workbook(out_filename)
    for i, sheet_name in enumerate(xl.sheetnames):
        i += 1
        worksheet = xl[sheet_name]
        
        worksheet.column_dimensions['A'].width = 36
        if i >=6:
            for row_idx in range(6, 21):
                worksheet.row_dimensions[row_idx].height = 40
            worksheet.row_dimensions[21].height = 50 
            worksheet.row_dimensions[6].height = 50 
            worksheet.row_dimensions[5].height = 125
            for c in range(1, 50):
                cell = worksheet.cell(row=6, column=c)
                cell.font = ft_column

            if i == 6:
                worksheet.column_dimensions['B'].width = 36
                worksheet.column_dimensions['C'].width = 36
                worksheet.column_dimensions['D'].width = 36
                worksheet.column_dimensions['E'].width = 36
                worksheet.column_dimensions['F'].width = 36
                worksheet.column_dimensions['G'].width = 36
            if i == 7:
                worksheet.column_dimensions['B'].width = 36
                worksheet.column_dimensions['C'].width = 36
                worksheet.column_dimensions['D'].width = 36
                worksheet.column_dimensions['E'].width = 36
                worksheet.column_dimensions['F'].width = 36
                worksheet.column_dimensions['G'].width = 36
            if i == 8:
                worksheet.row_dimensions[6].height = 60
                worksheet.column_dimensions['B'].width = 36
                worksheet.column_dimensions['C'].width = 36
                worksheet.column_dimensions['D'].width = 36
                worksheet.column_dimensions['E'].width = 36
                worksheet.column_dimensions['F'].width = 36
                worksheet.column_dimensions['G'].width = 36
                worksheet.column_dimensions['H'].width = 36


            
        else:
            worksheet.row_dimensions[5].height = 50
            for row_idx in range(6, 20):
                worksheet.row_dimensions[row_idx].height = 40
            if i == 1:
                worksheet.column_dimensions['D'].width = 34
                worksheet.column_dimensions['H'].width = 28
                worksheet.column_dimensions['I'].width = 22

        if i in [9, 10]:
            worksheet.row_dimensions[6].height = 100
            for c in range(2, 14):
                worksheet.column_dimensions[get_column_letter(c)].width = 25
            
            if i == 10:
                worksheet.column_dimensions['B'].width = 25



        if i in [11, 12, 13]:
            worksheet.row_dimensions[6].height = 100
            for c in range(2, 14):
                worksheet.column_dimensions[get_column_letter(c)].width = 22
                worksheet.row_dimensions[20].height = 50 
                if i in [11,12]:
                    worksheet.row_dimensions[6].height = 51
                elif i == 13:
                    worksheet.row_dimensions[6].height = 68
        
        if i in [14]:
            worksheet.row_dimensions[6].height = 113
            worksheet.column_dimensions['E'].width = 22
            worksheet.column_dimensions['K'].width = 22
            worksheet.column_dimensions['B'].width = 26
            worksheet.column_dimensions['B'].width = 23
            worksheet.column_dimensions['C'].width = 21
            worksheet.column_dimensions['D'].width = 21
            worksheet.column_dimensions['E'].width = 29
            worksheet.column_dimensions['F'].width = 21
            worksheet.column_dimensions['G'].width = 21
            worksheet.column_dimensions['H'].width = 21
            worksheet.column_dimensions['I'].width = 22
            worksheet.column_dimensions['J'].width = 21
            worksheet.column_dimensions['K'].width = 21
            worksheet.column_dimensions['L'].width = 40
            worksheet.row_dimensions[6].height = 73

        if i == 5:
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 40
            worksheet.column_dimensions['E'].width = 40
            worksheet.column_dimensions['F'].width = 40
            worksheet.column_dimensions['G'].width = 40
            worksheet.column_dimensions['H'].width = 40
            worksheet.row_dimensions[5].height = 145
        if i == 15:
            for c in range(2, 14):
                worksheet.column_dimensions[get_column_letter(c)].width = 26
            
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['I'].width = 33
            worksheet.row_dimensions[6].height = 75
            
        if i in [14, 15, 7, 6, 2]:
            for row_i in range(7, 22):
                worksheet.row_dimensions[row_i].height = 48
        
        if i in [10, 9]:
            for row_i in range(7, 22):
                worksheet.row_dimensions[row_i].height = 53
        
        # if i in [5]:
        #     for row_i in range(7, 22):
        #         worksheet.row_dimensions[row_i].height = 61.5
        
        
        if i in [15]:
            worksheet.row_dimensions[6].height = 85

        if i in list(range(1, 23)):
            # worksheet.column_dimensions['B'].width = 25
            # worksheet.column_dimensions['D'].width = 28
            # worksheet.column_dimensions['J'].width = 28
            worksheet.row_dimensions[20].height = 61.5
            worksheet.row_dimensions[21].height = 15
            worksheet.row_dimensions[22].height = 15
            worksheet.row_dimensions[23].height = 15
            worksheet.row_dimensions[24].height = 15
            worksheet.row_dimensions[25].height = 15
        
        if i in list(range(23, 100)):
            # worksheet.column_dimensions['B'].width = 25
            # worksheet.column_dimensions['D'].width = 28
            # worksheet.column_dimensions['J'].width = 28
            worksheet.row_dimensions[19].height = 61.5
            worksheet.row_dimensions[20].height = 15
            worksheet.row_dimensions[21].height = 15
            worksheet.row_dimensions[22].height = 15
            worksheet.row_dimensions[23].height = 15
            worksheet.row_dimensions[24].height = 15
            worksheet.row_dimensions[25].height = 15
        

        if i in [23, 24]:
            worksheet.row_dimensions[5].height = 257.25
            for letter_num in range(2,28):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21
        
        if  'Ер статистикаси' in sheet_name:
            worksheet.row_dimensions[5].height = 166.5
            for letter_num in range(2,5):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 33.29
        
        if  'Кредитга б' in sheet_name:
            worksheet.row_dimensions[5].height = 159.75
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 69.43
        
        if  'Харажатлар' in sheet_name:
            worksheet.row_dimensions[5].height = 122.25
            for letter_num in range(2,9):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 47.57
        
        if  '5.4. 2022 йилда ердан олинган д' in sheet_name:
            worksheet.row_dimensions[5].height = 176.25
            for letter_num in range(2,19):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 18.71
                if get_column_letter(letter_num) in ['F', 'G']:
                    worksheet.column_dimensions[get_column_letter(letter_num)].width = 24.86
        

        if  '4.7. Қандай қонунбу' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        
        if '4.4. Сизнингча, ерлардан унумли' in sheet_name:
            worksheet.row_dimensions[5].height = 168
            for letter_num in range(2,22):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 12.29
                if get_column_letter(letter_num) in ['J', 'K']:
                    worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.57
        

        if  '4.3. Фаолиятингизда қанд' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,30):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 12.29
        

        if  '4.2. Фаолиятингизга Қишлоқ хўжа' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,12):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
               
        if  '4.1. Ҳосилдорликни ошириш бўйич' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,12):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        

        if  '3.3. Қайси турдаги қишлоқ хўжал' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        
        if  '2.11. Сув таъминоти бил' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            worksheet.row_dimensions[6].height = 54.75
            for letter_num in range(2,12):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.86
        
        if  '2.10. Суғоришни' in sheet_name:
            worksheet.row_dimensions[5].height = 87.75
            worksheet.row_dimensions[6].height = 60.75
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 20.29
        

        if  '2.9. Ер майдонларини суғ' in sheet_name:
            worksheet.row_dimensions[5].height = 98.25
            worksheet.row_dimensions[6].height = 54
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.29
        

        if  '2.5. Ер олишда қандай' in sheet_name:
            worksheet.row_dimensions[5].height = 125
            for letter_num in range(2,14):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.29
        

        if  '1.9. (АГАР ҲОКИМ ЁРДАМЧИСИ БУНИ' in sheet_name:
            worksheet.row_dimensions[5].height = 86.25
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21.29
        

        if  '5.3. Ушбу ердан фойдаланиш уй х' in sheet_name:
            worksheet.row_dimensions[5].height = 61.5
            worksheet.row_dimensions[6].height = 49.5
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.86
        

        if  '5.1. Ушбу ер майдонларига эгали' in sheet_name:
            worksheet.row_dimensions[5].height = 68.25
            worksheet.row_dimensions[6].height = 49.5
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 20.71
        

        if  '4.5. Қўшимча молиявий ресурслар' in sheet_name:
            worksheet.row_dimensions[5].height = 71.25
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 21
        
        if  '2.8. Ажрати' in sheet_name:
            worksheet.row_dimensions[5].height = 114
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 18.57
        

        if  '2.7. Ер ажратилган пайтд' in sheet_name:
            worksheet.row_dimensions[5].height = 90.75
            for letter_num in range(2,10):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 17.14
        
        if  '2.4. Фойдаланаётган ер майдонин' in sheet_name:
            worksheet.row_dimensions[5].height = 72.75
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 20.86
        
        if  '2.3. Деҳқончилик қилиш, экинлар' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        

        if  '1.3. 2022 йилда ушбу ' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29

        
        if  '1.2. Респонде' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,8):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29
        


        if  '1.1.6. (БУНИ ЎҚИМАНГ) Респонден' in sheet_name:
            worksheet.row_dimensions[5].height = 50
            for letter_num in range(2,6):
                worksheet.column_dimensions[get_column_letter(letter_num)].width = 19.29

                    

        
            
        
        for i in range(5, 100):
            cell = worksheet.cell(row=i, column=1)
            cell.alignment = Alignment(horizontal='left', 
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=True,
                                        shrink_to_fit=True,
                                        indent=0) 
        

        


        
        

        
        
        
        xl.active = xl[sn]
        ws = xl.active
        
        

        
        
        

    xl.save(out_filename)
        # print(sheet_name)

    # wb = openpyxl.load_workbook(out_filename)
    # sheetNameList = wb.get_sheet_names()
    # for sheetName in sheetNameList:
    #     ws_active = wb.get_sheet_by_name(sheetName)
    #     cell = ws_active.cell(row=2, column=1)
    #     cell.value = get_rich_text(cell.value)
    #     if sheetName == 'Sheet5':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             cell.value = get_rich_text(cell.value, size=SIZE)
        
    #     if sheetName == 'Sheet6':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)

    #     if sheetName == 'Sheet7':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)
    #     if sheetName == 'Sheet9':
    #         for i in range(2, 9):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)
    #     if sheetName == 'Sheet10':
    #         for i in range(2, 15):
    #             cell = ws_active.cell(row=5, column=i)
    #             if not cell.value is None:
    #                 cell.value = get_rich_text(cell.value, size=SIZE)


    #     ws_active.sheet_properties.pageSetUpPr.fitToPage = True
    #     ws_active.page_setup.fitToHeight = True
    #     ws_active.page_setup.page_margins = True

        
    #     if sheetName in list(cells_.keys()):
    #         ft = ft_blue
    #         for cell_element in cells_[sheetName]:
    #             if sheetName =='Sheet1':
    #                 cell_element[1] = 7
    #             if sheetName =='Sheet5':
    #                 cell_element[1] = 3
    #                 ft = ft_red
    #             cell = ws_active.cell(row=cell_element[2], column=cell_element[1])
    #             cell.font = ft
    
    # wb.save(out_filename)



    

    print(f'{datetime.today()} - Starting generating REGIONAL REPORT file...')

