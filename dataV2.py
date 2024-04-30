

from datetime import datetime
from download import excel_to_pandas
from default import URL
from web import to_pdf
import pandas as pd 
from default import REPORT_COLS, outpath
import os 
import sys
import movecolumn as mc 
import numpy as np
from operator import itemgetter
import string
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

from utils.utils import get_rich_text, SIZE, FONT_NAME
import xlwings as xw
from xlwings.constants import PageBreak
import warnings
warnings.filterwarnings('ignore')

def convert_to_dict_of_lists(lst):
    result_dict = {}
    for item in lst:
        key = item[0]
        value = item[1]
        if key in result_dict:
            result_dict[key].append(value)
        else:
            result_dict[key] = [value]
    return result_dict

root = os.getcwd()
report_cols_list = [REPORT_COLS[key] for key in list(REPORT_COLS.keys())]
alphabet = dict.fromkeys(string.ascii_uppercase, 0)
alphabet = dict(zip(range(len(list(alphabet.keys()))), list(alphabet.keys())))


def to_excel(df, name):
    name = os.path.join(outpath, name)
    df.to_excel(name, index=False)







def generate_pdf():
    today = str(datetime.today().date()).replace('-', '_')
    db_filename = 'db_{today}.xlsx'.format(today=today)
    df = excel_to_pandas(URL, 'data/' + db_filename)
    return df

    

def report(df):
    mode ='w'
    prefix = ''
    today = str(datetime.today().date()).replace('-', '_')
    report_filename = 'report_analysis_{today}.xlsx'.format(today=today)
    # df=pd.read_excel(f'out\\db_{today}.xlsx')

    df['_submission_time'] = pd.to_datetime(df['_submission_time']).dt.date
    mask_eff = ~pd.isnull(df['7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?'])
    df = df[mask_eff]
    df_orig = df.copy()
    df.rename(columns={'Яшаш ҳудудингиз:': 'region', 'Яшаш тумани:': 'district', '_id': 'count'}, inplace=True)
    df = df[['region', 'district', 'count', '_submission_time']]

    
    dates = df_orig['_submission_time'].unique()
    df_out = pd.DataFrame()
    for i, date in enumerate(sorted(dates)):
        mask = df['_submission_time'] == date
        df_temp = df[mask]
        df_temp.drop('_submission_time', axis=1, inplace=True)
        df_temp = df_temp.groupby(by=['district'], as_index=False).count()
        df_temp.rename(columns={'region':'Ҳудуд',  'district': 'Туман', 'count': date}, inplace=True)
        total = df_temp[date].sum()
        df_temp.drop('Ҳудуд', axis=1, inplace=True)
        
        
        if i == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='outer', on=['Туман'])
    
    
    total_col = df_out[dates.tolist()].sum(axis=1)
    df_out['Жами']= total_col

    # sys.exit()



    # df_total = pd.DataFrame(columns=df_out.columns, data=[['Жами:', np.nan, np.nan, np.nan, np.nan, np.nan, np.nan , total]])
    df_total = df_out[dates].sum()
    total_all = sum(df_total.values.tolist())
    df_total = pd.DataFrame(columns=['Туман'] +df_total.index.values.tolist(), data=[['Жами:']+df_total[dates].values.tolist()])
    
    df_out = pd.concat([df_out, df_total], axis=0)
    df_out.reset_index(drop=True, inplace=True)
    df_out.at[df_out.index.values.tolist()[-1], 'Жами'] = total_all
    with pd.ExcelWriter(f'out_test/report/{report_filename}', mode=mode, engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}districts')

    # print(df_out)
    ############################################################################ Вилоятлар    
    df_orig.rename(columns={'Яшаш ҳудудингиз:': 'region', 'Яшаш тумани:': 'district', '_id': 'count'}, inplace=True)
    df_orig = df_orig[['region', 'district', 'count', '_submission_time']]

    # df_orig = df_orig.groupby(by=['region'], as_index=False).count()[['region', 'count']]
    # df_orig.rename(columns={'region':'Ҳудуд', 'count': 'Уй хўжаликлар сони'}, inplace=True)
    df_out = pd.DataFrame()
   
    for k, date in enumerate(dates):
        mask = df_orig['_submission_time'] == date
        df_temp = df_orig[mask]
        df_temp = df_temp.groupby(by=['region'], as_index=False).count()[['region', 'count']]
        total = df_temp['count'].sum()
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        df_total = pd.DataFrame(columns=df_temp.columns, data=[['Жами:', total]])
        
        df_temp.rename(columns={'region':'Ҳудуд', 'count': date}, inplace=True)
        
        if k == 0:
            df_out = df_temp
        else:
            df_out = pd.merge(df_out, df_temp, how='outer', on=['Ҳудуд'])
    df_out['Жами'] = df_out[dates].sum(axis=0)
    total_row = df_out.sum()[dates]
    total_row["Ҳудуд"] = 'Жами:'
    total_row = total_row.reset_index().T.reset_index(drop=True)
    total_row.rename(dict(zip(total_row.columns.values.tolist(), total_row.iloc[0].values.tolist())), inplace=True, axis=1)
    total_row.drop(0, axis=0, inplace=True)
    df_out = pd.concat([df_out, total_row], axis=0)
    columns = ["Ҳудуд"] + sorted(dates.tolist()) + ['Жами']
    df_out = df_out[columns]
    df_out['Жами'] = df_out[dates].sum(axis=1)

    df_plan = pd.read_excel('plan.xlsx')
    df_out = pd.merge(df_out, df_plan, on='Ҳудуд', how='left')
    df_out['Ҳолати, %'] = (df_out['Жами'].div(df_out['Режа'])*100).round(1)
    with pd.ExcelWriter(f'out_test/report/{report_filename}', mode='a', engine='openpyxl') as writer:
        df_out.to_excel(writer, sheet_name=f'{prefix}regions')
    print(f'{datetime.today()} - Starting generating REPORT file...')
    return report_filename


def highlight(x):
    print(x)
    return x

def freq_table(df):
    
    # if 'out' not in df_filename:
    #     df_filename = os.path.join(root, 'out', df_filename)
    # df = pd.read_excel(df_filename)
    mask_eff = ~pd.isnull(df['7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?'])
    df = df[mask_eff]
    df.dropna(axis=1, inplace=True, how='all')
    df_orig = df.copy()
    
    
    today = str(datetime.today().date()).replace('-', '_')


    columns = ['52. Маълумотингиз қандай?',
               '53. Ҳозирги пайтда асосий фаолиятингиз қандай?',
               '54. Қайси соҳада ишлайсиз?',
               '55. Илтимос, оилангизнинг жорий молиявий аҳволини тасвирлаб беринг?',
                '7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?',
               '7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?',
               '8. Вилоятингиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?',
               '9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?',
               '10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?',
               '11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?',
               '12. Сизнингча, атрофингиздаги одамлар ҳозир қандай кайфиятда: кўтаринки, хотиржам ёки тушкун (хавотирли)?',
               '13. “Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:',
               '14. “Мамлакатимизда олиб борилаётган ислоҳотлар тўғри йўлда кетмоқда”, мазкур фикрга:', 
               '16. “Ҳукумат фуқаролар билан очиқ мулоқотда бўлмоқда ва уларнинг муаммоларига ўз вақтида жавоб қайтармоқда" Мазкур фикрга...:',
               '25. Президент Шавкат Мирзиёев ўз лавозимида фикрингиз бўйича қандай фаолият кўрсатаётганлигини 7 баллик шкалада баҳоланг?',
               '33.  Сизнингча, (вилоят ҳокими) вилоят/республика ҳокими сифатида қандай ишлаяпти?',
               '36. Сизнингча, туманингиз ҳокими ўз лавозимида қандай ишламоқда?']
    columns = ['Яшаш ҳудудингиз:'] + columns

    df = df[columns]
    df.reset_index(drop=True, inplace=True)

    columns = df.columns
    df_out = pd.DataFrame()

    filename_out = f'out_test/freq/ecopol_freq_table_{today}.xlsx'
    df_big = pd.DataFrame()

    town_vil = [ 'Шаҳар', 'Қишлоқ', 'Аёл', 'Эркак', '18-35', '36-49', '50-999','all',]
    for i, c in enumerate(list(reversed(df.columns.values))):

        
        df_temp = pd.DataFrame()
        for j, tv in enumerate(town_vil):
            if tv in ['Шаҳар', 'Қишлоқ']:
                mask = df_orig['1. САВОЛ БЕРМАЙ, Қаерда (Шаҳар ёки қишлоқда) яшашини киритинг?'] == tv
                df = df_orig[mask]
            elif tv in ['18-35', '36-49', '50-999']:
                btm_age, top_age = tv.split('-')
                btm_age, top_age = int(btm_age), int(top_age)
                if btm_age == 18:
                    btm_age = 17
                btm_mask = df_orig['6. Ёшингиз?'] >= btm_age
                top_mask = df_orig['6. Ёшингиз?'] <= top_age
                # mask = ~(btm_mask & top_mask)
                df = df_orig[btm_mask][top_mask]
                df = df.dropna(subset='6. Ёшингиз?', axis=0)
                
            elif tv in ['Аёл', 'Эркак']:
                mask = df_orig['2. САВОЛ БЕРМАЙ, респондентларнинг жинсини аниқланг:'] == tv
                df = df_orig[mask]
            else:
                df = df_orig.copy()
            
            ctab = pd.crosstab(index = df.index, columns = df[c])
            ctab = ctab.sum()

            ctab_pct = (ctab/ctab.values.sum()).round(decimals=4)
            
            tab = pd.concat([ctab, ctab_pct], axis=1)

            cols = ['Frequency', 'Percentage (%)', c]
            tab[c] = tab.index
            
            df_out = pd.DataFrame(columns=cols, data=tab.values)
            df_out = mc.MoveTo1(df_out, c)

            if not df[c].dtype == 'object':
                df_out['Average score'] = df_out['Percentage (%)']/100 * df_out[c]
                df_out['Average score'] = df_out['Average score'].round(decimals=3)
            else:
                df_out['Average score'] = [np.nan] * len(df_out['Percentage (%)'])
            
            fillna_cols = df_out.columns.values.tolist()
            
            df_out[['Percentage (%)', 'Average score']] = df_out[['Percentage (%)', 'Average score']] * 100

            
            freq_sum = df_out['Frequency'].sum()
            pct_sum =df_out['Percentage (%)'].sum()
            avg_sum = df_out['Average score'].sum()
            if avg_sum == 0:
                avg_sum = np.nan
            df_out['Frequency'] = df_out['Frequency'].astype(float).astype('Int64')
            df_total = pd.DataFrame(columns=df_out.columns, data=[['Total:', freq_sum,\
                                                                pct_sum,
                                                                 avg_sum]])

            df_out = pd.concat([df_out, df_total], axis=0)
            
            

            
            # Add space after tables
            df_out.rename(columns={'Frequency': f'{tv}_Frequency', 'Percentage (%)': f'{tv}_Percentage (%)', 'Average score': f'{tv}_Average score'}, inplace=True)
            if j == 0:
                df_temp = df_out
            else:
                if tv == 'all' or tv in ['Аёл'] or tv == '18-35':
                    space = pd.DataFrame(np.nan, index=list(range(len(df_out))), columns=['A', 'B'])
                    # space = df_out[df_out.columns.values[:2]].fill(0)
                    df_out.reset_index(drop=True, inplace=True)
                    df_out = pd.concat([space, df_out], axis=1)
                
         
                df_temp = pd.merge(df_temp, df_out, how='outer', on=c)
                fillna_cols = df_temp.columns.values.tolist()
                fillna_cols = [fc for fc in fillna_cols if fc != 'Average score' and fc != c and fc not in ['A', 'B'] and 'A_' not in fc and 'B_' not in fc]
                df_temp[fillna_cols] = df_temp[fillna_cols].fillna(0)
            

        df_space = pd.DataFrame(columns=df_temp.columns.values.tolist(), data=[[np.nan]*len(df_temp.columns),
                                                                                [np.nan]*len(df_temp.columns)])
        
        ### move the Total row to the last
        col_first = df_temp.columns.values[0]
        mask = df_temp[col_first] == 'Total:'
        df_temp = pd.concat([df_temp[~mask], df_temp[mask]], axis=0)

        ## fill nan values with zeroes
        # df_temp.fillna(0, inplace=True)

        ## add space between question blocks
        df_temp = pd.concat([df_temp, df_space], axis=0)

        

        # columns = [c] + ['Frequency_x', 'Frequency_y', 'Percentage (%)_x',  'Percentage (%)_y', 'Average score_x', 'Average score_y', 'A', 'B', 'Frequency', 'Percentage (%)', 'Average score']
        # print(df_temp.columns.values)
        # sys.exit()
        columns = [c] + ['Шаҳар_Frequency', 'Қишлоқ_Frequency', 'Шаҳар_Percentage (%)',  'Қишлоқ_Percentage (%)', 'Шаҳар_Average score', 'Қишлоқ_Average score', 'A_x', 'B_x', \
                          'Аёл_Frequency', 'Эркак_Frequency', 'Аёл_Percentage (%)', 'Эркак_Percentage (%)', 'Аёл_Average score', 'Эркак_Average score', 'A_y', 'B_y',
                          '18-35_Frequency', '36-49_Frequency', '50-999_Frequency', '18-35_Percentage (%)', '36-49_Percentage (%)', '50-999_Percentage (%)', '18-35_Average score', '36-49_Average score', '50-999_Average score', 'A', 'B',
                          'all_Frequency', 'all_Percentage (%)', 'all_Average score']
        


        

        micolumns = pd.MultiIndex.from_tuples(
                                            [(None, c),
                                             ("Frequency", "Шаҳар"), (None, "Қишлоқ"), 
                                             ("Percentage (%)", "Шаҳар"), (None, "Қишлоқ"),
                                             ("Average score", "Шаҳар"), (None, "Қишлоқ"),
                                             (None, None),
                                             (None, None),
                                             ("Frequency", "Аёл"), (None, "Эркак"), 
                                             ("Percentage (%)", "Аёл"), (None, "Эркак"),
                                             ("Average score", "Аёл"), (None, "Эркак"),
                                             (None, None),
                                             (None, None),
                                             ("Frequency", '18-35'),(None, '36-49'),(None, '50+'), 
                                             ("Percentage (%)", '18-35'),(None, '36-49'),(None, '50+'),
                                             ("Average score", "18-35"),(None, "36-49"),(None, "50+"),
                                             (None, None),
                                             (None, None),
                                             ('Across all data', "Frequency"), 
                                             (None, "Percentage (%)"),
                                             (None, "Average score")],)
        
        # columns = [c] + ['all_Frequency', 'all_Percentage (%)', 'all_Average score']
   
        # micolumns = pd.MultiIndex.from_tuples(
        #                                     [(None, c),
        #                                      ('Across all data', "Frequency"), 
        #                                      (None, "Percentage (%)"),
        #                                      (None, "Average score")
        #                                      ],)

        df_temp = df_temp[columns]
        df_temp = pd.DataFrame(columns=micolumns, data=df_temp.values)

    
        
        
 
        if i == 0:
            df_big = df_temp
        else:
            
            df_temp = df_temp.reset_index(drop=True).T.reset_index().T
            df_temp.rename(columns=dict(zip(df_temp.columns.values.tolist(), df_big.columns.values.tolist())), inplace=True)
            if i == 1:
                columns = df_big.columns.values.tolist()
                df_big = df_big.reset_index(drop=True).T.reset_index().T
                df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), columns)), inplace=True)
            
            df_big = pd.concat([df_temp, df_big], axis=0)

    
        

    new_columns = [np.nan] * len(df_big.columns.values)
    new_columns[0] = 'EcoPol Frequency table'
    new_columns[6] = today.replace('_', '-')
    df_big.rename(columns=dict(zip(df_big.columns.values.tolist(), new_columns)), inplace=True)
    
    with pd.ExcelWriter(filename_out, mode='w', engine='openpyxl') as writer:
        df_big.to_excel(writer, index=False)



    print(f'{datetime.today()} - Starting generating FREQ TABLE file...')
    return filename_out
    

def report_by_region(df):
    
    # today = str(datetime.today().date()).replace('-', '_')
    # df_filename = f'out\\db_{today}.xlsx'
    # df = pd.read_excel(df_filename)
    today = str(datetime.today().date()).replace('-', '_')
    out_filename = f'out_test/regional/regional_{today}.xlsx'
    
    # cols = ['7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?',
    #         '7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?',
    #         '11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?',
    #         '9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?',
    #         '12. Сизнингча, атрофингиздаги одамлар ҳозир қандай кайфиятда: кўтаринки, хотиржам ёки тушкун (хавотирли)?',
    #         '13. “Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:',
    #         '16. “Ҳукумат фуқаролар билан очиқ мулоқотда бўлмоқда ва уларнинг муаммоларига ўз вақтида жавоб қайтармоқда" Мазкур фикрга...:',
    #         '14. “Мамлакатимизда олиб борилаётган ислоҳотлар тўғри йўлда кетмоқда”, мазкур фикрга:',
    #         '26. Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?',
    #         '27. Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?',
    #         '30. Сизнингча, Шавкат Мирзиёев мамлакатдаги вазиятни яхши томонга ўзгартира оладими ёки йўқми?']

    cols = {
        'multiple': ['15. Давлат томонидан қуйида келтирилган соҳалардан қай бирида амалга оширилаётган ишларни маъқуллайсиз?',
                     '28. Илтимос, Шавкат Мирзиёевнинг Президентлик даврида эришган асосий ютуқларини номлаб беринг?',
                     '29. Илтимос, Шавкат Мирзиёевнинг Президентлик давридаги асосий камчиликларини номлаб беринг?',
                     '51. Сиз яшаб турган ҳудуддаги энг асосий муаммони кўрсатинг:',
                     ' балл'],
        'multicolumn': ['Умумий вазият яхшиланмоқдами ёки деярли ўзгармаяптими?', 'Мамлакат ва ислоҳотларга ишонч',
                        'Амалдаги Президентга ишонч', 'Ҳукумат ва Парламент раҳбарларига муносабат','Сиёсий партиялар раҳбарларига муносабат',
                        'Вилоят ҳокимларига муносабат',
                        'Туман/шаҳар ҳокимларига муносабат',
                        'Парламент сайловларига муносабат',
                        'Маълумотларнинг асосий манбаси',
                        'Президент сайловига муносабат']
    }
    
    region_col = 'Яшаш ҳудудингиз:'
    df_out = pd.DataFrame()
    merge_columns = {}
    dfs = []
    sheet_names = []
    counter = 0
    for type_idx, (type_, columns_) in enumerate(cols.items()):
        for sheet_idx, col in enumerate(columns_):
            if type_ == 'single':
                ctab = pd.crosstab(index = df[region_col], columns = df[col])
                ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col], normalize='index') * 100).round(decimals=1)
                ctab = ctab_pct
                total_by_region = ctab.sum(axis=1)
                total_all = total_by_region.sum()
                total_by_region.name='Жами'
                # ctab = pd.concat([ctab, total_by_answers], axis=0)
                total_by_region= pd.DataFrame(total_by_region)
                # ctab = pd.concat([total_by_region['Жами'], ctab], axis=1)
                
                total_by_answers = ctab.sum(axis=0)
                total_by_answers = total_by_answers.reset_index().T
                total_by_answers.rename(dict(zip(total_by_answers.columns.values.tolist(), total_by_answers.iloc[0].values.tolist())), inplace=True, axis=1)
            
                total_by_answers = total_by_answers.tail(-1)
                total_by_answers.index = ['Жами:']
                ctab = pd.concat([ctab, total_by_answers], axis=0)
                
                
                # ctab = pd.concat([ctab['Жами'], ctab_pct], axis=1)
                ctab.drop(index='Жами:', inplace=True)
                total_by_answers_pct = (total_by_answers.div(total_all)).astype(float).multiply(100).round(decimals=1)
                
                # total_by_answers_pct['Жами'] = None
                total_by_answers_pct.index=['Республика бўйича']
                ctab = pd.concat([ctab, total_by_answers_pct], axis=0)
                ctab = pd.concat([ctab, total_by_answers], axis=0)

            elif type_ == 'multiple':
                if col == ' балл':
                    columns = [
                        '25. Президент Шавкат Мирзиёев ўз лавозимида фикрингиз бўйича қандай фаолият кўрсатаётганлигини 7 баллик шкалада баҳоланг?',
                        '33.  Сизнингча, (вилоят ҳокими) вилоят/республика ҳокими сифатида қандай ишлаяпти?',
                        "36. Сизнингча, туманингиз ҳокими ўз лавозимида қандай ишламоқда?",
                        "40. Ўзбекистон Республикаси Олий Мажлис Қонунчилик палатасининг амалдаги таркибининг фаолиятини қандай баҳолайсиз?",
                        "7.1. Мамлакатдаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?",
                        "8. Вилоятингиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?",
                        "10. Маҳаллангиздаги умумий вазиятдан қониқиш даражангизни 7 баллик шкалада баҳоланг?"
                    ]
                                    
                    
                else:
                    columns = [c for c in df.columns.values if col in c][1:]
                    drop_by_key = ['Ҳеч бири',
                                   "Бошқа",
                                   "Жавоб бериш",
                                   "Ҳеч қандай",
                                   "Телефон алоқаси билан боғлиқ муаммо"]
                    for key in drop_by_key:
                        columns = [c for c in columns if not key in c]
                    
                dfs_multiple = []
                for col_ in columns: 
                    ctab = pd.crosstab(index = df[region_col], columns = df[col_])
                    ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col_], normalize='index', margins=True) * 100).round(decimals=1)
                    index = ctab_pct.index.values.tolist()
                    index[-1] = 'Республика бўйича'
                    ctab_pct.index = index
                    ctab_pct = pd.concat([ctab_pct.tail(1), ctab_pct.head(-1)], axis=0)
                    ctab = ctab_pct
                    
                    # total_by_answers_pct['Жами'] = None
                    col_name = col_.replace(col+'/', '')
                    
                    if col == ' балл':
                        col_name = col_name.split(' ', maxsplit=1)[1]

                        column_values = pd.Series(ctab.columns.values.tolist(), index=ctab.columns)
                        ctab = ctab/100
                        ctab = ctab.multiply(column_values, axis=1)
                        ctab['Average score'] = ctab.sum(axis=1)
                        ctab = ctab['Average score'].to_frame().round(1)
                        ctab.rename(columns={'Average score': 1}, inplace=True)
                    ctab.rename(columns={1:col_name}, inplace=True)
                    dfs_multiple.append(ctab[col_name].to_frame())
                
                ctab = pd.concat(dfs_multiple, axis=1)

            elif type_ == 'multicolumn':
                
                # columns = [c for c in df.columns.values if col in c]
                if col == 'Умумий вазият яхшиланмоқдами ёки деярли ўзгармаяптими?':
                    columns = ['7.2. Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?',
                               '9. Сизнингча, вилоятдаги умумий вазият яхшиланмоқда, ёмонлашмоқда ёки деярли ўзгармаяптими?',
                               '11. Сизнингча, сиз яшаётган маҳаллада умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки ҳеч нарса ўзгармаяптими?']
                    
                if col == 'Мамлакат ва ислоҳотларга ишонч':
                    columns = ['13. “Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:',
                               '14. “Мамлакатимизда олиб борилаётган ислоҳотлар тўғри йўлда кетмоқда”, мазкур фикрга:',
                               '16. “Ҳукумат фуқаролар билан очиқ мулоқотда бўлмоқда ва уларнинг муаммоларига ўз вақтида жавоб қайтармоқда" Мазкур фикрга...:']
                    
                if col == 'Амалдаги Президентга ишонч':
                    columns = ['26. Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?',
                               '27. Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?',
                               '30. Сизнингча, Шавкат Мирзиёев мамлакатдаги вазиятни яхши томонга ўзгартира оладими ёки йўқми?',]
                if col == 'Ҳукумат ва Парламент раҳбарларига муносабат':
                    columns = ['18. Абдулла Арипов, Ўзбекистон Республикаси Бош вазири?',
                               '19. Танзила Норбоева, Ўзбекистон Республикаси Олий Мажлиси Сенати раиси?',
                               '20. Нуриддин Исмоилов, Ўзбекистон Республикаси Олий Мажлиси Қонунчилик палатасининг спикери?']
                
                if col == 'Сиёсий партиялар раҳбарларига муносабат':
                    columns = ['21. Шавкат Мирзиёев, Ўзбекистон Либерал-демократик партиясидан президентликка номзод?',
                               '22. Роба Маҳмудова, Ўзбекистон «Адолат»  демократик партиясидан президентликка номзод?',
                               '23. Улуғбек Иноятов, Ўзбекистон Халқ демократик партиясидан президентликка номзод?',
                               '24. Абдушукур Ҳамзаев, Ўзбекистон Экологик партиядан президентликка номзод?']
                
                if col == 'Вилоят ҳокимларига муносабат':
                    columns = ['32. Айтингчи, Сиз вилоят ҳокимига ишонасизми ёки ишонмайсизми?',
                               '34. Сизнингча сиз яшаётган вилоятга янги ҳоким керакми ёки амалдаги ҳоким қолгани яхшими?',
                               '38. Сиз қандай фикрдасиз – қайси бири яхши: вилоят ҳокимини ҳозиргидек тайинлаган маъқулми ёки уни овоз бериш орқали сайлаганми?',
                               ]
                
                if col == 'Туман/шаҳар ҳокимларига муносабат':
                    columns = ['35. Сиз туманингиз ҳокимига ишонасизми ёки ишонмайсизми?',
                               '37. Сизнингча туманга янги ҳоким керакми ёки амалдаги ҳоким қолгани яхшими?',
                               '39. Сиз қандай фикрдасиз – қайси бири яхши: туман ҳокимини ҳозиргидек тайинлаган маъқулми ёки уни овоз бериш орқали сайлаганми?',
                               ]
                
                if col == 'Парламент сайловларига муносабат':
                    columns = ['41. Агар Олий Мажлис Қонунчилик палатаси сайловлари кейинги якшанба куни бўлиб ўтса, сиз қайси партияга овоз берган бўлардингиз?',
                               '44. Ҳар қандай шароитда ҳам қайси партияларга овоз бермайсиз?',
                               ]
                
                if col == 'Маълумотларнинг асосий манбаси':
                    columns = ['49. Мамлакат ҳаётидаги қизиқарли воқеалар ҳақида асосан қаердан маълумот оласиз?',
                               '50. Ўзингиз учун қайси ижтимоий тармоқни асосий деб ҳисоблайсиз?',
                               ]
                if col == 'Президент сайловига муносабат':
                    columns = ['46. Одатда, сайловга борасизми?',
                               '47. Сиз нима деб ўйлайсиз, Ўзбекистон Республикаси Президенти сайловларида қатнаша оласизми ёки қатнаша олмайсизми?',
                               '45. Ўзбекистон Республикаси президенти сайлови қачон ўтказилишини биласизми?'
                               ]
                
                if '“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:' in columns:
                    col_idx = columns.index('“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:')
                    columns[col_idx] = '“Мен Ўзбекистон иқтисодиёти ривожланишига ишонаман”,  мазкур фикрга:'
                new_columns = [c.split(' ', maxsplit=1)[1].replace('...', '') for c in columns]
                
                df.rename(columns=dict(zip(columns, new_columns)), inplace=True)
                columns = new_columns
                dfs_multiple = []
                for col_ in columns:
                    
                    ctab = pd.crosstab(index = df[region_col], columns = df[col_])
                    ctab_pct = (pd.crosstab(index = df[region_col], columns = df[col_], normalize='index', margins=True) * 100).round(decimals=1)
                    index = ctab_pct.index.values.tolist()
                    index[-1] = 'Республика бўйича'
                    ctab_pct.index = index
                    ctab_pct = pd.concat([ctab_pct.tail(1), ctab_pct.head(-1)], axis=0)
                    ctab = ctab_pct
                    col_name = col_.replace(col+'/', '')
                    if col == 'Умумий вазият яхшиланмоқдами ёки деярли ўзгармаяптими?':
                        if col_ == 'Сизнингча, мамлакатдаги умумий вазият яхшиланмоқдами, ёмонлашмоқдами ёки деярли ўзгармаяптими?':
                            target_cols = ['Яхшиланмоқда', 'Деярли ўзгармаяпти']
                        else:
                            target_cols = ['Яхшиланмоқда', 'Умуман ўзгармаяпти']
                        
                        data = {}
                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                        

                    elif col == 'Мамлакат ва ислоҳотларга ишонч':
                        target_cols = ['Тўлиқ қўшиламан', 'Қисман қўшиламан']
                        data = {}
                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                        
                        
                    elif col == 'Амалдаги Президентга ишонч':
                        if col_ == 'Ўтган икки ой мобайнида Шавкат Мирзиёевга Ўзбекистон Президенти сифатида муносабатингиз ўзгардими? Ва, агар ўзгарган бўлса, у қайси томонга - яхшиланди ёки ёмонлашди?':
                            target_cols = ['Анча яхшиланди', 'Озроқ яхшиланди', 'Ҳеч нарса ўзгармади']
                        elif col_ == 'Шавкат Мирзиёевга ишонасизми ёки ишонмайсизми?':
                            target_cols = ['Тўлиқ ишонаман', 'Ишонаман']
                        elif col_== 'Сизнингча, Шавкат Мирзиёев мамлакатдаги вазиятни яхши томонга ўзгартира оладими ёки йўқми?':
                            target_cols = ['Аниқ ўзгартира олади', 'Ўзгартириши мумкин', 'Ўзгартириши қийин']
                        

                        data = {}
                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                    
                    elif col == 'Ҳукумат ва Парламент раҳбарларига муносабат':
                        target_cols = ['Буни билмайман', 'Биламан ва ишонч билан ижобий муносабатда бўламан', 'Биламан, лекин муносабатни баҳолай олмайман']
                        data = {}
                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()

                    
                    elif col == 'Сиёсий партиялар раҳбарларига муносабат':
                        target_cols = ['Буни билмайман', 'Биламан ва ишонч билан ижобий муносабатда бўламан', 'Биламан, лекин муносабатни баҳолай олмайман']
                        data = {}
                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                        
                    elif col == 'Вилоят ҳокимларига муносабат':
                        if col_ == 'Айтингчи, Сиз вилоят ҳокимига ишонасизми ёки ишонмайсизми?':
                            target_cols = ['Тўлиқ ишонаман', 'Қисман ишонаман', 'Умуман ишонмайман']
                        elif col_ == 'Сизнингча сиз яшаётган вилоятга янги ҳоким керакми ёки амалдаги ҳоким қолгани яхшими?':
                            target_cols = ['Янги ҳоким керак', 'Амалдаги қолгани яхши', 'Менга фарқи йўқ']
                        elif col_== 'Сиз қандай фикрдасиз – қайси бири яхши: вилоят ҳокимини ҳозиргидек тайинлаган маъқулми ёки уни овоз бериш орқали сайлаганми?':
                            target_cols = ['Тайинланса яхши', 'Сайланса яхши', 'Менга фарқи йўқ']
                        data = {}
                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                    
                    elif col == 'Туман/шаҳар ҳокимларига муносабат':
                        if col_ == 'Сиз туманингиз ҳокимига ишонасизми ёки ишонмайсизми?':
                            target_cols = ['Тўлиқ ишонаман', 'Ишонаман', 'Ишонмайман']
                        elif col_ == 'Сизнингча туманга янги ҳоким керакми ёки амалдаги ҳоким қолгани яхшими?':
                            target_cols = ['Янги ҳоким керак', 'Амалдаги қолгани яхши', 'Менга фарқи йўқ']
                        elif col_== 'Сиз қандай фикрдасиз – қайси бири яхши: туман ҳокимини ҳозиргидек тайинлаган маъқулми ёки уни овоз бериш орқали сайлаганми?':
                            target_cols = ['Тайинланса яхши', 'Сайланса яхши', 'Менга фарқи йўқ']
                        data = {}

                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                    
                    elif col == 'Парламент сайловларига муносабат':
                        target_cols = ['УзЛиДеп', '“Миллий тикланиш”', 'Халқ демократик партияси', '“Адолат”', 'Экология партияси']
                        data = {}

                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                        
                    elif col == 'Маълумотларнинг асосий манбаси':
                        if col_ == 'Мамлакат ҳаётидаги қизиқарли воқеалар ҳақида асосан қаердан маълумот оласиз?':
                            target_cols = ['Телевидение', 'Интернет', 'Танишлар, дўстлар', 'Босма нашрлар (Газета, журналлар)', 'Радио']
                        elif col_ == 'Ўзингиз учун қайси ижтимоий тармоқни асосий деб ҳисоблайсиз?':
                            target_cols = ['Телеграм', 'Фейсбук', 'Инстаграм', 'Ютьюб','Тик-ток', 'Ижтимоий тармоқларда аккаунтим йўқ/ижтимоий тармоқларга кирмайман']
                        
                        data = {}

                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()
                    
                    elif col == 'Президент сайловига муносабат':
                        if col_ == 'Одатда, сайловга борасизми?':
                            target_cols = ['Мен деярли барча сайловларга бораман', 'Баъзан сайловларга бораман', 'Бормайман']
                        elif col_ == 'Сиз нима деб ўйлайсиз, Ўзбекистон Республикаси Президенти сайловларида қатнаша оласизми ёки қатнаша олмайсизми?':
                            target_cols = ['Албатта қатнашаман', 'Аниқ айта олмайман', 'Ҳозирча аниқ эмас, сайловга яқин қарор қиламан', 'Йўқ']
                        elif col_ == 'Ўзбекистон Республикаси президенти сайлови қачон ўтказилишини биласизми?':
                            target_cols = ['Билмайман/Жавоб беришга қийналаман', 'Аниқ санасини айтди']
                        
                        data = {}

                        for ti, target_col in enumerate(target_cols):
                            data[(col_, target_col)] = ctab[target_col].values.tolist()

                    ctab = pd.DataFrame(data, index=ctab.index)
                    dfs_multiple.append(ctab)
                
                ctab = pd.concat(dfs_multiple, axis=1)


            
            # ctab = (pd.crosstab(index = df[region_col], columns = df[col], normalize='index') * 100).round(decimals=1)
            ctab[col] = ctab.index 
            ctab = mc.MoveTo1(ctab, col)
            ctab = ctab.reset_index(drop=True).T.reset_index().T      
            ctab.reset_index(drop=True, inplace=True)
            
            # ctab.style.background_gradient(cmap='coolwarm').set_precision(2)
            # ctab.show()
            merge_columns[ctab.iloc[0][0]] = len(ctab.columns)

            dfs.append(ctab)
            sn = f'Sheet{counter+1}'
            if counter == 0:
                with pd.ExcelWriter(out_filename, mode='w', engine='openpyxl') as writer:
                    ctab.T.reset_index().T.to_excel(writer, index=False,header=None, sheet_name=sn)
            else:
                with pd.ExcelWriter(out_filename, mode='a', engine='openpyxl', if_sheet_exists='error') as writer:
                    ctab.T.reset_index().T.to_excel(writer, index=False,header=None, sheet_name=sn)

            counter +=1
            sheet_names.append(sn)
            
            # removing question nums from headers
            if col[0].isnumeric():
                col = col.split(' ', maxsplit=1)[1] 
            workbook = openpyxl.load_workbook(filename=out_filename)
            worksheet = workbook[sn]
            worksheet.sheet_view.zoomScale = 55
            ft_default = Font(size=SIZE, name=FONT_NAME)
            ft_blue = Font(color="0070c0", bold=True, size=SIZE, name=FONT_NAME)
            ft_red = Font(color="ff0000", bold=True, size=SIZE, name=FONT_NAME)
            ft_column = Font(color='1f4e78', bold=True, size=SIZE, name=FONT_NAME)
            ft_tiny = Font(color='1f4e78', italic=True, size=14, name=FONT_NAME)
            edge_col = len(ctab.columns)
            for c in range(edge_col):
                cell = worksheet.cell(row=1, column=c+1)
                cell.value = None
            cell = worksheet.cell(row=1, column=1)
            if col == ' балл':
                col = 'Қониқиш даражаси'

            if col == 'Умумий вазият яхшиланмоқдами ёки деярли ўзгармаяптими?':
                col = 'Умумий вазият яхшиланмоқдами ёки деярли ўзгармаяптими?'
            
            if col == 'Илтимос, Шавкат Мирзиёевнинг Президентлик давридаги асосий камчиликларини номлаб беринг?':
                col = 'Илтимос, 2017-2022 йилларда давлат сиёсатида асосий камчиликларни номлаб беринг?'


            cell.value = col
          
            
            cell.font = ft_column
            cell.alignment = Alignment(horizontal='center', 
                                vertical='center',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0) 
            worksheet.row_dimensions[2].height = 50
            worksheet.insert_rows(1)
            worksheet.insert_rows(3)
            worksheet.insert_rows(3)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=edge_col)
            
            cell = worksheet.cell(row=5, column=1)
            cell.value = None
            worksheet.column_dimensions['A'].width = 34
            # cell = worksheet.cell(row=1, column=edge_col)
            # cell.value = f'{counter}-жадвал'

            cell = worksheet.cell(row=4, column=edge_col)
            cell.value = f'(фоизда)'
            cell.alignment = Alignment(horizontal='right', 
                                vertical='bottom',
                                text_rotation=0,
                                wrap_text=False,
                                shrink_to_fit=False,
                                indent=0) 

            if counter == 5:
                cell.value = f'(7 баллик ўлчов бирлигида, 1- энг ёмон, 7 энг юқори)'
                cell.alignment = Alignment(horizontal='right', 
                                vertical='bottom',
                                text_rotation=0,
                                wrap_text=False,
                                shrink_to_fit=False,
                                indent=0) 
            
            cell.font = ft_tiny

            if type_ == 'multicolumn':
                worksheet.row_dimensions[6].height = 60
                columns = ctab.iloc[0].tolist()[1:]
                all_cols = ctab.iloc[1].values.tolist()[1:]
                columns = list(zip(columns, all_cols))
                columns = convert_to_dict_of_lists(columns)
                key_count = 1
                for _, col_ in enumerate(list(columns.keys())):
                    target_cols = columns[col_]
                    
                    for ti, target_col in enumerate(target_cols):
                        key_count += 1
                        if ti == len(target_cols)-1:
                            worksheet.merge_cells(start_row=5, start_column=key_count-len(target_cols)+1, end_row=5, end_column=key_count)
                            worksheet[f'{alphabet[key_count-len(target_cols)]}5'] = col_
                            break
                        
                        cell = worksheet.cell(row=5, column=key_count)
                        cell.value = None
                        
                    
                    # worksheet[f'{alphabet[key_count-1]}5'] = col_

                    cell = worksheet.cell(row=6, column=key_count)
                    cell.font = ft_column

                worksheet.row_dimensions[6].height = 60
            
            for j in range(5, 22):
                alignment = Alignment(horizontal='general', 
                                vertical='bottom',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0) 
                cell = worksheet.cell(row=j, column=1)
                cell.alignment = alignment
                worksheet.row_dimensions[j].height = 30
                if j == 5:
                    worksheet.row_dimensions[j].height = 120
                    
                for k in range(1, edge_col+1):
                    cell = worksheet.cell(row=j, column=k)
                    cell.font = ft_default
                    if k == 1 and j == 6:
                        cell.font = ft_column
                        continue
                    if k != 1:
                        alignment.horizontal = 'center' 
                        alignment.vertical = 'center' 
                    if j == 5:
                        if len(ctab.columns) > 11:
                            if not col in ['Маълумотларнинг асосий манбаси', 'Сиёсий партиялар раҳбарларига муносабат']:
                                alignment.text_rotation = 90
                                worksheet.row_dimensions[j].height = 190
                            worksheet.column_dimensions[alphabet[k]].width = 13
                        else:
                            worksheet.column_dimensions[alphabet[k]].width = 20
                            

                        cell.font = ft_column
                    
                    
                    else:  
                        
                        rep_cell = worksheet.cell(row=6, column=1)
                        if rep_cell.value == 'Республика бўйича':
                            rep_cell.font = ft_column
                            rep_vals = [ (worksheet.cell(row=6, column=c_).value, c_)  for c_ in range(1, edge_col)][1:]
                            rep_vals = [v[1] for v in rep_vals if v[0] >=50]
                        elif worksheet.cell(row=7, column=1).value == 'Республика бўйича':
                            rep_cell = worksheet.cell(row=7, column=1)
                            rep_cell.font = ft_column
                            rep_vals = [(worksheet.cell(row=7, column=c_).value, c_)  for c_ in range(1, edge_col)][1:]
                            rep_vals = [v[1] for v in rep_vals if v[0] >=50]
                            
                        if not cell.value is None:
                            if not isinstance(cell.value, str):
                                if not col in ['Қониқиш даражаси']:
                                    if k in rep_vals:
                                        if cell.value >= 50:
                                            cell.font = ft_blue
                                        else:
                                            cell.font = ft_red

                                else:
                                    if cell.value >= 6:
                                        cell.font = ft_blue
                                    elif cell.value <=2:
                                        cell.font = ft_red
                        

                        

                    cell.alignment = alignment


            border = Border(
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000')
                        ) 

            cell_range = worksheet[f'A5:{alphabet[edge_col-1]}{len(ctab)+4}']
            for row in cell_range:
                for cell in row:
                    cell.border = border
            worksheet.print_options.horizontalCentered = True
            worksheet.print_options.verticalCentered = True
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3
            workbook.save(out_filename)
            
    xl = openpyxl.load_workbook(out_filename)
    for i, sheet_name in enumerate(xl.sheetnames):
        i += 1
        worksheet = xl[sheet_name]
            
        if i >=6:
            worksheet.row_dimensions[21].height = 50 
            worksheet.row_dimensions[6].height = 50 
            worksheet.row_dimensions[5].height = 125
            for c in range(1, 50):
                cell = worksheet.cell(row=6, column=c)
                cell.font = ft_column

            if i == 6:
                worksheet.column_dimensions['B'].width = 25
                worksheet.column_dimensions['D'].width = 25
                worksheet.column_dimensions['F'].width = 25
            if i == 7:
                worksheet.column_dimensions['G'].width = 28
            if i == 8:
                worksheet.row_dimensions[6].height = 60
                worksheet.column_dimensions['H'].width = 22
                worksheet.column_dimensions['I'].width = 22


            
        else:
            worksheet.row_dimensions[20].height = 50 
            if i == 1:
                worksheet.column_dimensions['D'].width = 34
                worksheet.column_dimensions['H'].width = 28
                worksheet.column_dimensions['I'].width = 22

        if i in [9, 10]:
            worksheet.row_dimensions[6].height = 100
            for c in range(2, 14):
                worksheet.column_dimensions[alphabet[c]].width = 25
            
            if i == 10:
                worksheet.column_dimensions['B'].width = 25



        if i in [11, 12, 13]:
            worksheet.row_dimensions[6].height = 100
            for c in range(2, 14):
                worksheet.column_dimensions[alphabet[c]].width = 22
                worksheet.row_dimensions[20].height = 50 
                if i in [11,12]:
                    worksheet.row_dimensions[6].height = 51
                elif i == 13:
                    worksheet.row_dimensions[6].height = 68
        
        if i in [14]:
            worksheet.row_dimensions[6].height = 113
            worksheet.column_dimensions['E'].width = 22
            worksheet.column_dimensions['K'].width = 22
            worksheet.column_dimensions['B'].width = 26
            worksheet.column_dimensions['B'].width = 23
            worksheet.column_dimensions['C'].width = 21
            worksheet.column_dimensions['D'].width = 21
            worksheet.column_dimensions['E'].width = 29
            worksheet.column_dimensions['F'].width = 21
            worksheet.column_dimensions['G'].width = 21
            worksheet.column_dimensions['H'].width = 21
            worksheet.column_dimensions['I'].width = 22
            worksheet.column_dimensions['J'].width = 21
            worksheet.column_dimensions['K'].width = 21
            worksheet.column_dimensions['L'].width = 40
            worksheet.row_dimensions[6].height = 73

        if i == 5:
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 40
            worksheet.column_dimensions['D'].width = 40
            worksheet.column_dimensions['E'].width = 40
            worksheet.column_dimensions['F'].width = 40
            worksheet.column_dimensions['G'].width = 40
            worksheet.column_dimensions['H'].width = 40
            worksheet.row_dimensions[5].height = 145
        if i == 15:
            for c in range(2, 14):
                worksheet.column_dimensions[alphabet[c]].width = 26
            
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['I'].width = 33
            worksheet.row_dimensions[6].height = 75
        
        for i in range(5, 100):
            cell = worksheet.cell(row=i, column=1)
            cell.alignment = Alignment(horizontal='left', 
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=True,
                                        shrink_to_fit=True,
                                        indent=0) 
     
        
        
        

        
        
        

    xl.save(out_filename)
        # print(sheet_name)




    

    print(f'{datetime.today()} - Starting generating REGIONAL REPORT file...')

